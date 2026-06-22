"""관세청 'HS부호 단위별 품목명' 공공데이터 → trade/data/hs_names.tsv 변환
(사용자 2026-06-22 '무실적 코드까지 전부 이름').

data.go.kr '관세청_HS부호 단위별 품목명'(15130660) 파일(CSV/Excel)을 받아,
HS코드 열 + 한글품목명 열만 뽑아 `code<TAB>한글명` TSV 로 저장한다. 헤더명이
판마다 달라도 자동 탐지(코드열=값이 대부분 숫자 / 한글명열=한글 비중 최다).

실행:
    .venv/bin/python -m trade.scripts.build_hs_names <받은파일.csv|.xlsx>
    .venv/bin/python -m trade.scripts.build_hs_names <파일> --out <경로>
    .venv/bin/python -m trade.scripts.build_hs_names <파일> --code-col HS부호 --name-col 한글품목명

graceful: 열 자동탐지 실패 시 --code-col/--name-col 로 지정. .xlsx 는 openpyxl
있을 때만(없으면 CSV 로 받아 재실행). 기존 hs_names.tsv 는 덮어씀.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys

_DEFAULT_OUT = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                            "data", "hs_names.tsv")
_HANGUL = re.compile(r"[가-힣]")


def _read_blocks(path: str) -> list[list[list[str]]]:
    """파일 → [블록(헤더+행)]. .xlsx = **전 시트**(관세청 파일은 HS2/4/6/8/10단위
    시트 분리, 사용자 2026-06-22), 각 시트가 1블록. .csv/.tsv = 단일 블록."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl 미설치 — data.go.kr 에서 CSV 로 받아 재실행하세요.")
        wb = load_workbook(path, read_only=True, data_only=True)
        blocks = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = [["" if c is None else str(c) for c in row]
                    for row in ws.iter_rows(values_only=True)]
            if rows:
                blocks.append(rows)
        return blocks
    # CSV/TSV — 인코딩 폴백(utf-8-sig → cp949, 정부 파일 cp949 흔함)
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                delim = "\t" if sample.count("\t") > sample.count(",") else ","
                return [[list(r) for r in csv.reader(f, delimiter=delim)]]
        except (UnicodeDecodeError, UnicodeError):
            continue
    sys.exit(f"인코딩 판독 실패: {path}")


def _digit_ratio(s: str) -> float:
    s = (s or "").strip()
    if not s:
        return 0.0
    return sum(c.isdigit() for c in s) / len(s)


def _detect_cols(rows: list[list[str]]) -> tuple[int, int]:
    """(코드열 idx, 한글명열 idx) 자동탐지. 헤더명 힌트 우선, 없으면 데이터 통계."""
    if not rows:
        sys.exit("빈 파일")
    header = [(_h or "").strip() for _h in rows[0]]
    code_i = name_i = -1
    for i, h in enumerate(header):
        hl = h.replace(" ", "")
        if code_i < 0 and ("HS" in hl.upper() or "부호" in hl or "코드" in hl):
            code_i = i
        if name_i < 0 and (("한글" in hl and "명" in hl) or hl == "품목명"
                           or ("국문" in hl and "명" in hl)):
            name_i = i
    # 데이터 통계 폴백(헤더 힌트 실패 시): 샘플 행에서 숫자비율 최대=코드, 한글 최다=명
    body = rows[1:200] or rows
    ncol = max((len(r) for r in body), default=0)
    if code_i < 0:
        code_i = max(range(ncol),
                     key=lambda c: sum(_digit_ratio(r[c]) for r in body if c < len(r)),
                     default=0)
    if name_i < 0:
        name_i = max((c for c in range(ncol) if c != code_i),
                     key=lambda c: sum(len(_HANGUL.findall(r[c])) for r in body if c < len(r)),
                     default=0)
    return code_i, name_i


def build(path: str, out: str = _DEFAULT_OUT,
          code_col: str | None = None, name_col: str | None = None) -> int:
    seen: dict[str, str] = {}
    for rows in _read_blocks(path):          # 시트(블록)별 — 헤더·열 매핑 각자
        if not rows:
            continue
        header = [(_h or "").strip() for _h in rows[0]]
        if code_col and code_col in header:
            code_i = header.index(code_col)
        else:
            code_i, _auto_n = _detect_cols(rows)
        if name_col and name_col in header:
            name_i = header.index(name_col)
        else:
            _auto_c, name_i = _detect_cols(rows)
        print(f"열 매핑: 코드='{header[code_i] if code_i < len(header) else code_i}' "
              f"명='{header[name_i] if name_i < len(header) else name_i}' ({len(rows)-1}행)")
        for r in rows[1:]:
            if code_i >= len(r) or name_i >= len(r):
                continue
            code = re.sub(r"\D", "", r[code_i] or "")
            name = (r[name_i] or "").strip()
            if not code or not name or not _HANGUL.search(name):
                continue
            seen[code] = name                # 동일 코드 최신만(마지막 승)
    if not seen:
        sys.exit("추출 0행 — --code-col/--name-col 로 열을 지정해 재실행하세요.")
    # 버전 헤더 — 파일명의 발효일(YYYYMMDD) + 빌드시각 + 행수. catalog_guard 가
    # 읽어 갱신 추적·연 1회 신규본 넛지(사용자 2026-06-22 '연계표처럼'). 로더는
    # '#' 줄을 무시(탭 없음)하므로 무해.
    import datetime as _dt
    m = re.search(r"(20\d{6})", os.path.basename(path))
    eff = m.group(1) if m else ""
    built = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        f.write(f"# effective={eff} built={built} rows={len(seen)} "
                f"src={os.path.basename(path)}\n")
        for code in sorted(seen):
            f.write(f"{code}\t{seen[code]}\n")
    print(f"✅ {len(seen):,}개 HS 한글품목명 (발효 {eff or '미상'}) → {out}")
    return len(seen)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("file", help="관세청 HS부호 단위별 품목명 파일(.csv/.xlsx)")
    ap.add_argument("--out", default=_DEFAULT_OUT)
    ap.add_argument("--code-col", default=None, help="코드 열 헤더명(자동탐지 실패 시)")
    ap.add_argument("--name-col", default=None, help="한글명 열 헤더명")
    a = ap.parse_args(argv)
    build(a.file, a.out, a.code_col, a.name_col)
    return 0


if __name__ == "__main__":
    sys.exit(main())
