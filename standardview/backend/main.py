#!/usr/bin/env python3
# backend/main.py  —  DART 재무분석기 FastAPI 백엔드
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
import asyncio, concurrent.futures, subprocess, shutil, tempfile, sqlite3
from dotenv import load_dotenv
import os, json, re, threading
import requests, zipfile, xml.etree.ElementTree as ET, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── 환경변수 로드 ──────────────────────────────────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

# ── SQLite DB ─────────────────────────────────────────────────────────────────
DB_PATH    = Path(__file__).parent.parent / "data" / "standard_view.db"
BACKUP_DIR = Path(__file__).parent.parent / "backups"
_OLD_DB    = Path(__file__).parent / "insights.db"

def _db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def _init_db():
    # DB 디렉터리 생성
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    # 구 DB 마이그레이션 (backend/insights.db -> data/standard_view.db)
    if not DB_PATH.exists() and _OLD_DB.exists():
        shutil.copy2(str(_OLD_DB), str(DB_PATH))
    with _db() as conn:

        conn.execute("""
            CREATE TABLE IF NOT EXISTS macro_cache (
                indicator_id TEXT PRIMARY KEY,
                value_json   TEXT,
                fetched_at   TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS macro_news_cache (
                cache_key  TEXT PRIMARY KEY,
                value_json TEXT,
                fetched_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS insights (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at   TEXT,
                author       TEXT,
                company_name TEXT,
                sector       TEXT,
                source_type  TEXT,
                source_title TEXT,
                source_url   TEXT,
                key_content  TEXT,
                user_insight TEXT,
                impact_level TEXT,
                next_actions TEXT,
                tags         TEXT
            )
        """)
        conn.commit()

_init_db()

# ── Constants ──────────────────────────────────────────────────────────────────
API_KEY  = os.getenv("DART_API_KEY", "")
CORP_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
FIN_URL  = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
_cur_year = datetime.now().year
YEARS    = [str(_cur_year - i) for i in range(5)]  # 최근 5개년 자동 탐색

CO_COLS  = ["3182F6", "FF6B35", "7C3AED"]
LV_COLOR = {"good": "#05C072", "note": "#F59E0B", "warn": "#F97316", "danger": "#F04452"}
LV_XL    = {"good": "05C072",  "note": "F59E0B",  "warn": "F97316",  "danger": "F04452"}
LV_XL_BG = {"good": "E8FBF3",  "note": "FFFBEB",  "warn": "FFF4ED",  "danger": "FFF0F1"}
LV_ICON  = {"good": "✓", "note": "!", "warn": "△", "danger": "✕"}

OPINION_KEYS = [
    ("수익성", "영업이익률"),   ("수익성", "순이익률"),
    ("수익성", "매출총이익률"), ("효율성", "ROE"),
    ("효율성", "ROA"),          ("안정성", "부채비율"),
    ("안정성", "유동비율"),     ("안정성", "이자보상배율"),
    ("안정성", "이익잉여금 비율"),
    ("성장성", "매출 성장률"),  ("성장성", "영업이익 성장률"),
    ("성장성", "순이익 성장률"),("총평",   "종합 의견"),
]

INCOME_KEYS  = ["매출","매출원가","매출총이익","판관비","영업이익",
                "금융수익","금융비용","세전이익","법인세비용","당기순이익","EPS"]
BALANCE_KEYS = ["유동자산","비유동자산","자산총계",
                "유동부채","비유동부채","부채총계","이익잉여금","자본총계"]
RATIO_KEYS   = ["영업이익률","순이익률","ROE","ROA","부채비율","유동비율","이자보상배율"]
SECTIONS     = [
    ("손익계산서",  INCOME_KEYS,  False),
    ("재무상태표",  BALANCE_KEYS, False),
    ("수익성 지표", RATIO_KEYS,   True),
]
EPS_KEYS   = {"EPS"}
XRATE_KEYS = {"이자보상배율"}

# ── CODE_MAP (account_id → canonical) ─────────────────────────────────────────
CODE_MAP: dict = {
    "ifrs-full_Revenue":                                "매출",
    "ifrs-full_RevenueFromContractsWithCustomers":      "매출",
    "dart_Revenue":                                     "매출",
    "ifrs-full_CostOfSales":                            "매출원가",
    "dart_CostOfSales":                                 "매출원가",
    "ifrs-full_GrossProfit":                            "매출총이익",
    "ifrs-full_SellingGeneralAndAdministrativeExpense": "판관비",
    "dart_SellingGeneralAdministrativeExpenses":        "판관비",
    "ifrs-full_AdministrativeExpense":                  "판관비",
    "dart_OperatingIncomeLoss":                         "영업이익",
    "ifrs-full_ProfitLossFromOperatingActivities":      "영업이익",
    "ifrs-full_FinanceIncome":                          "금융수익",
    "ifrs-full_FinanceCosts":                           "금융비용",
    "ifrs-full_ProfitLossBeforeTax":                    "세전이익",
    "ifrs-full_IncomeTaxExpenseContinuingOperations":   "법인세비용",
    "ifrs-full_ProfitLoss":                             "당기순이익",
    "dart_ProfitLoss":                                  "당기순이익",
    "us-gaap_NetIncomeLoss":                            "당기순이익",
    "ifrs-full_BasicEarningsLossPerShare":              "EPS",
    "ifrs-full_BasicEarningsPerShare":                  "EPS",
    "dart_BasicEarningsLossPerShare":                   "EPS",
    "ifrs-full_CurrentAssets":                          "유동자산",
    "ifrs-full_NoncurrentAssets":                       "비유동자산",
    "ifrs-full_Assets":                                 "자산총계",
    "us-gaap_Assets":                                   "자산총계",
    "ifrs-full_CurrentLiabilities":                     "유동부채",
    "ifrs-full_NoncurrentLiabilities":                  "비유동부채",
    "ifrs-full_Liabilities":                            "부채총계",
    "ifrs-full_RetainedEarnings":                       "이익잉여금",
    "ifrs-full_Equity":                                 "자본총계",
    "us-gaap_StockholdersEquity":                       "자본총계",
    "ifrs-full_EquityAttributableToOwnersOfParent":     "자본총계",
}

# ── NAME_MAP (account_nm → canonical, fallback) ───────────────────────────────
NAME_MAP: dict = {
    "매출액": "매출", "수익(매출액)": "매출", "영업수익": "매출",
    "매출": "매출", "도급공사수익": "매출", "분양수익": "매출",
    "이자수익": "매출",
    "매출원가": "매출원가", "도급공사원가": "매출원가",
    "매출총이익": "매출총이익", "매출총이익(손실)": "매출총이익",
    "매출총손익": "매출총이익",
    "판매비와관리비": "판관비", "판매비및관리비": "판관비",
    "판매비와 관리비": "판관비", "영업비용": "판관비",
    "영업이익": "영업이익", "영업이익(손실)": "영업이익", "영업손익": "영업이익",
    "금융수익": "금융수익", "이자및배당금수익": "금융수익",
    "금융비용": "금융비용", "이자비용": "금융비용", "금융원가": "금융비용",
    "법인세비용차감전순이익": "세전이익",
    "법인세비용차감전순이익(손실)": "세전이익",
    "법인세차감전순이익": "세전이익",
    "법인세비용차감전순손익": "세전이익",
    "법인세비용": "법인세비용", "법인세수익": "법인세비용",
    "당기순이익": "당기순이익", "당기순이익(손실)": "당기순이익",
    "분기순이익": "당기순이익", "반기순이익": "당기순이익",
    "당기순손익": "당기순이익",
    "기본주당순이익": "EPS", "기본주당이익": "EPS",
    "기본주당순이익(손실)": "EPS", "보통주기본주당이익(손실)": "EPS",
    "유동자산": "유동자산",
    "비유동자산": "비유동자산", "고정자산": "비유동자산",
    "자산총계": "자산총계", "자산 합계": "자산총계",
    "유동부채": "유동부채",
    "비유동부채": "비유동부채", "고정부채": "비유동부채",
    "부채총계": "부채총계", "부채 합계": "부채총계",
    "이익잉여금": "이익잉여금", "이익잉여금(결손금)": "이익잉여금",
    "미처분이익잉여금": "이익잉여금",
    "자본총계": "자본총계", "자본 합계": "자본총계",
}

# ── Utilities ──────────────────────────────────────────────────────────────────
def _parse(s, as_float=False):
    try:
        s = str(s).replace(",", "").strip()
        if s and s not in ("-", ""):
            return float(s) if as_float else int(float(s))
    except Exception:
        pass
    return None


def fmt_num(n, key=""):
    if n is None: return "N/A"
    if key in EPS_KEYS: return f"{int(n):,}원"
    v = n / 1e8
    if abs(v) >= 10000: return f"{v/10000:,.1f}조"
    return f"{v:,.1f}억"


def fmt_ratio(v, key=""):
    if v is None: return "N/A"
    if key in XRATE_KEYS: return f"{v:.2f}x"
    return f"{v:.1f}%"


def fmt_val(v, key="", is_ratio=False):
    return fmt_ratio(v, key) if is_ratio else fmt_num(v, key)


def yoy_str(cur, prev):
    if cur is None or prev is None or prev == 0: return ""
    r = (cur - prev) / abs(prev) * 100
    if r > 0:   return f"▲{abs(r):.1f}%"
    elif r < 0: return f"▼{abs(r):.1f}%"
    return "─"


# ── DART Data Layer ────────────────────────────────────────────────────────────
class DART:
    def __init__(self):
        self.corps: dict = {}
        self.ready = False

    def load(self):
        r = requests.get(CORP_URL, params={"crtfc_key": API_KEY}, timeout=30)
        r.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            with z.open(z.namelist()[0]) as f:
                root = ET.parse(f).getroot()
        self.corps = {}
        for item in root.findall("list"):
            nm = item.findtext("corp_name", "").strip()
            if not nm: continue
            self.corps.setdefault(nm, []).append({
                "corp_code":  item.findtext("corp_code", "").strip(),
                "stock_code": item.findtext("stock_code", "").strip(),
            })
        self.ready = True

    def search(self, q: str) -> list:
        if not self.ready: return []
        ql = q.lower()
        return [
            {"corp_name": nm, **c}
            for nm, corps in self.corps.items()
            if ql in nm.lower()
            for c in corps
        ][:30]

    def financials(self, corp_code: str, year: str):
        for fs in ("CFS", "OFS"):
            try:
                r = requests.get(FIN_URL, params={
                    "crtfc_key": API_KEY, "corp_code": corp_code,
                    "bsns_year": year, "reprt_code": "11011", "fs_div": fs,
                }, timeout=30)
                d = r.json()
                if d.get("status") == "000" and d.get("list"):
                    return d["list"], fs
            except Exception:
                pass
        return None, None

    def extract(self, items: list) -> dict:
        if not items: return {}
        res: dict = {}
        for item in items:
            acct_id   = (item.get("account_id") or "").strip()
            acct_nm   = (item.get("account_nm") or "").strip()
            canonical = CODE_MAP.get(acct_id) or NAME_MAP.get(acct_nm)
            if not canonical: continue
            v = _parse(item.get("thstrm_amount", ""), as_float=(canonical in EPS_KEYS))
            if v is None: continue
            if canonical not in res or abs(v) > abs(res[canonical]):
                res[canonical] = v
        return res

    @staticmethod
    def calc_ratios(d: dict) -> dict:
        rev   = d.get("매출")      or 0
        op    = d.get("영업이익")
        net   = d.get("당기순이익")
        asset = d.get("자산총계")  or 0
        eq    = d.get("자본총계")  or 0
        dbt   = d.get("부채총계")  or 0
        ca    = d.get("유동자산") or 0
        cl    = d.get("유동부채") or 0
        fc    = d.get("금융비용") or 0
        gp    = d.get("매출총이익") or 0
        er    = d.get("이익잉여금") or 0
        return {
            "영업이익률":    op  / rev   * 100 if op  is not None and rev   else None,
            "순이익률":      net / rev   * 100 if net is not None and rev   else None,
            "ROE":          net / eq    * 100 if net is not None and eq    else None,
            "ROA":          net / asset * 100 if net is not None and asset else None,
            "부채비율":      dbt / eq    * 100 if eq                        else None,
            "유동비율":      ca  / cl    * 100 if cl                        else None,
            "이자보상배율":  op  / fc          if op  is not None and fc    else None,
            "매출총이익률":  gp  / rev   * 100 if gp  and rev               else None,
            "이익잉여금 비율": er / asset * 100 if er and asset              else None,
        }


# ── Claude CLI ─────────────────────────────────────────────────────────────────
CLAUDE_PROMPT = """\
당신은 한국 주식 투자 전문 재무 분석가입니다.
아래 기업 재무 데이터를 분석하고, 투자자 관점의 의견을 JSON으로만 출력하세요.
JSON 이외의 텍스트(설명, 마크다운 등)는 절대 출력하지 마세요.

{fin_table}

다음 13개 항목을 분석하여 정확히 아래 JSON 형식으로 출력하세요:
{{
  "영업이익률":      {{"value":"최신값(예:11.9%)", "text":"투자관점 의견 25자이내", "level":"good|note|warn|danger"}},
  "순이익률":        {{"value":"...", "text":"...", "level":"..."}},
  "매출총이익률":    {{"value":"...", "text":"...", "level":"..."}},
  "ROE":            {{"value":"...", "text":"...", "level":"..."}},
  "ROA":            {{"value":"...", "text":"...", "level":"..."}},
  "부채비율":        {{"value":"...", "text":"...", "level":"..."}},
  "유동비율":        {{"value":"...", "text":"...", "level":"..."}},
  "이자보상배율":    {{"value":"...(예:5.2x)", "text":"...", "level":"..."}},
  "이익잉여금 비율": {{"value":"...", "text":"...", "level":"..."}},
  "매출 성장률":     {{"value":"▲X.X% 또는 ▼X.X%", "text":"...", "level":"..."}},
  "영업이익 성장률": {{"value":"...", "text":"...", "level":"..."}},
  "순이익 성장률":   {{"value":"...", "text":"...", "level":"..."}},
  "종합 의견":       {{"value":"A|B|C|D", "text":"종합 등급 한 줄 평가", "level":"..."}}
}}
level 기준 — good:양호/강점, note:보통/중립, warn:주의/약점, danger:위험/심각"""


def check_claude_cli() -> bool:
    """Patched 2026-05-19 (higgack/standardview deploy):
    Claude Code CLI -> Gemini API. Reuses GOOGLE_API_KEY env var
    (same key as NOAH stock-bot). Function name preserved for call-site
    backward compatibility throughout main.py."""
    return bool(os.environ.get("GOOGLE_API_KEY"))


_gemini_client = None
_GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")


def _log_sv_usage(prompt_tok, output_tok, cost_krw, endpoint=""):
    import json as _json, os as _os
    from datetime import datetime as _dt
    try:
        path = Path(_os.path.expanduser("~/standardview/sv_usage.jsonl"))
        path.parent.mkdir(parents=True, exist_ok=True)
        now = _dt.now()
        rec = {
            "ts": now.isoformat(timespec="seconds"),
            "date": now.date().isoformat(),
            "month": now.date().isoformat()[:7],
            "prompt_tok": prompt_tok,
            "output_tok": output_tok,
            "cost_krw": round(cost_krw, 4),
            "endpoint": endpoint,
        }
        with path.open("a", encoding="utf-8") as f:
            f.write(_json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        pass


def call_claude_cli(prompt: str, timeout: int = 120) -> str | None:
    """Patched 2026-05-19/20: Claude Code CLI -> google-genai SDK +
    token usage cost logging to ~/standardview/sv_usage.jsonl."""
    api_key = os.environ.get("GOOGLE_API_KEY")
    if not api_key or not prompt or not prompt.strip():
        return None
    global _gemini_client
    if _gemini_client is None:
        try:
            from google import genai
            _gemini_client = genai.Client(api_key=api_key)
        except Exception:
            return None
    try:
        resp = _gemini_client.models.generate_content(
            model=_GEMINI_MODEL,
            contents=prompt,
        )
        text = (resp.text or "").strip()
        try:
            um = getattr(resp, "usage_metadata", None)
            if um is not None:
                pt = getattr(um, "prompt_token_count", 0) or 0
                ot = getattr(um, "candidates_token_count", 0) or 0
                # gemini-2.5-flash real pricing: $0.30 in / $2.50 out per 1M
                # (was a stale 0.075/0.30 gemini-1.5 constant → undercounted ~7.5x).
                cost = (pt * 0.30 + ot * 2.50) / 1e6 * 1330.0
                _log_sv_usage(pt, ot, cost)
        except Exception:
            pass
        return text or None
    except Exception:
        return None


def _extract_json(text: str) -> dict | list | None:
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    # try to extract JSON array too
    m = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.DOTALL)
    if m:
        try: return json.loads(m.group(1))
        except Exception: pass
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try: return json.loads(m.group(1))
        except Exception: pass
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        try: return json.loads(text[start:end+1])
        except Exception: pass
    return None


def _normalize_opinions(raw: dict) -> dict:
    result = {}
    for cat, lbl in OPINION_KEYS:
        item = raw.get(lbl, {})
        if not isinstance(item, dict): item = {}
        result[lbl] = {
            "category": cat,
            "value":    str(item.get("value", "N/A")),
            "text":     str(item.get("text",  "분석 데이터 없음")),
            "level":    item.get("level", "note")
                        if item.get("level") in LV_COLOR else "note",
        }
    return result


def _build_fin_table(corp_name: str, yearly: dict) -> str:
    avail = [yr for yr in YEARS if yearly.get(yr)]
    if not avail:
        return f"회사명: {corp_name}\n데이터 없음"
    lines = [f"회사명: {corp_name}  |  분석 기준: {avail[-1]}년 결산\n"]

    def row(label, vals, suffix="억원", fmt_fn=None):
        parts = []
        for i, (yr, v) in enumerate(zip(avail, vals)):
            if v is None:
                parts.append(f"{yr}: N/A")
                continue
            s = fmt_fn(v) if fmt_fn else f"{v/1e8:,.1f}{suffix}"
            if i > 0 and vals[i-1] is not None and vals[i-1] != 0:
                r = (v - vals[i-1]) / abs(vals[i-1]) * 100
                arrow = f"▲{r:.1f}%" if r > 0 else f"▼{abs(r):.1f}%"
                parts.append(f"{yr}: {s} ({arrow})")
            else:
                parts.append(f"{yr}: {s}")
        return f"  {label:<18} " + "  ".join(parts)

    lines.append("[ 손익계산서 ] (단위: 억원)")
    for key in INCOME_KEYS:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        if key == "EPS":
            lines.append(row(key, vals, "원", lambda v: f"{int(v):,}원"))
        else:
            lines.append(row(key, vals))

    lines.append("\n[ 재무상태표 ] (단위: 억원)")
    for key in BALANCE_KEYS:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        lines.append(row(key, vals))

    lines.append("\n[ 수익성·안정성 지표 ]")
    for key in RATIO_KEYS + ["매출총이익률", "이익잉여금 비율"]:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        fn = (lambda v: f"{v:.2f}x") if key in XRATE_KEYS else (lambda v: f"{v:.1f}%")
        lines.append(row(key, vals, "", fn))

    return "\n".join(lines)


def generate_opinions_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = CLAUDE_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    return _normalize_opinions(raw)




# ── Deal Intelligence Prompts ──────────────────────────────────────────────────
DEAL_SIGNAL_PROMPT = """당신은 VC/M&A/PE/IPO 전문 딜 어드바이저입니다.
아래 기업의 재무 데이터를 기반으로 현재 딜 투자 가치를 평가하고 JSON으로만 출력하세요.
JSON 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "signal":       "Strong Buy|Watch|Caution|Pass 중 하나",
  "signal_level": "good|note|warn|danger",
  "summary":      "핵심 한 줄 평가 (20자 이내)",
  "reasons":      ["근거 1 (25자 이내)", "근거 2", "근거 3"],
  "timing":       "지금이 딜 적기인 이유 또는 아닌 이유 (40자 이내)",
  "deal_types":   ["VC", "M&A", "PE", "IPO", "CFO" 중 해당 항목들]
}}
signal_level: Strong Buy→good, Watch→note, Caution→warn, Pass→danger"""

INDUSTRY_RADAR_PROMPT = """당신은 산업 분석 전문 투자 애널리스트입니다.
아래 기업의 재무 데이터를 기반으로 산업 포지셔닝을 분석하고 JSON으로만 출력하세요.
JSON 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "industry":         "추정 업종/산업명",
  "market_dynamics":  "시장 동학 한 줄 설명 (40자 이내)",
  "moat":             "경쟁 해자 한 줄 설명 (40자 이내)",
  "growth_vectors":   ["성장 동력 1 (25자 이내)", "성장 동력 2", "성장 동력 3"],
  "threats":          ["주요 위협 1 (25자 이내)", "주요 위협 2"],
  "positioning":      "Leader|Challenger|Niche|Vulnerable 중 하나",
  "positioning_text": "산업 내 포지셔닝 한 줄 평 (30자 이내)"
}}"""

DEAL_MEMO_PROMPT = """당신은 투자은행 시니어 어드바이저입니다.
아래 기업의 재무 데이터를 기반으로 {deal_type} 관점의 Deal Memo를 Markdown으로 작성하세요.
Markdown 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

반드시 아래 구조를 따르세요:

# Deal Memo: {corp_name}

**딜 유형:** {deal_type} | **작성일:** {date} | **기밀 등급:** Confidential

---

## 1. Executive Summary

## 2. Why Now?

## 3. Company Overview

## 4. Industry View

## 5. Financial Checkpoints

## 6. Deal Angle

## 7. Key Due Diligence Questions

## 8. Risks

## 9. Next Actions"""


# ── Deal Intelligence: Mock Functions ─────────────────────────────────────────
def _deal_signal_mock(corp_name: str, yearly: dict) -> dict:
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default=None)
    if not latest_yr:
        return {"signal": "Pass", "signal_level": "danger",
                "summary": "재무 데이터 없음",
                "reasons": ["DART 데이터를 불러올 수 없습니다"],
                "timing": "데이터 확인 후 재시도 필요", "deal_types": []}
    d = yearly[latest_yr]
    op_margin = d.get("영업이익률")
    roe       = d.get("ROE")
    prev_yr   = str(int(latest_yr) - 1)
    rev = d.get("매출"); prev_rev = yearly.get(prev_yr, {}).get("매출")
    rev_growth = (rev - prev_rev) / abs(prev_rev) * 100 if rev and prev_rev and prev_rev != 0 else None

    score = 0; reasons = []
    if op_margin is not None:
        if op_margin >= 15:  score += 2; reasons.append(f"영업이익률 {op_margin:.1f}% — 수익성 우수")
        elif op_margin >= 5: score += 1; reasons.append(f"영업이익률 {op_margin:.1f}% — 안정적")
        else:                score -= 1; reasons.append(f"영업이익률 {op_margin:.1f}% — 수익성 개선 필요")
    if roe is not None:
        if roe >= 15:  score += 2; reasons.append(f"ROE {roe:.1f}% — 자본 효율 우수")
        elif roe >= 8: score += 1; reasons.append(f"ROE {roe:.1f}% — 보통 수준")
        else:          score -= 1; reasons.append(f"ROE {roe:.1f}% — 자본 효율 낮음")
    if rev_growth is not None:
        if rev_growth >= 20:  score += 2; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 고성장")
        elif rev_growth >= 5: score += 1; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 성장세")
        else:                 score -= 1; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 성장 둔화")
    if not reasons:
        reasons.append(f"{latest_yr}년 재무 데이터 기반 평가")

    if score >= 4:   signal, level, timing, dtypes = "Strong Buy", "good",   "재무지표 전반 양호, 딜 검토 적기",       ["VC", "M&A", "PE", "IPO"]
    elif score >= 2: signal, level, timing, dtypes = "Watch",      "note",   "일부 지표 긍정적, 추가 실사 권고",       ["M&A", "PE"]
    elif score >= 0: signal, level, timing, dtypes = "Caution",    "warn",   "리스크 존재, 신중한 접근 필요",          ["PE"]
    else:            signal, level, timing, dtypes = "Pass",       "danger", "재무 구조 개선 후 재검토 권고",          []
    return {"signal": signal, "signal_level": level,
            "summary": f"{corp_name} — {signal}",
            "reasons": reasons[:3], "timing": timing, "deal_types": dtypes}


def _industry_radar_mock(corp_name: str, yearly: dict) -> dict:
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default=None)
    d = yearly.get(latest_yr, {}) if latest_yr else {}
    op_margin = d.get("영업이익률")
    if op_margin and op_margin >= 15:   pos, pos_txt = "Leader",     "높은 수익성으로 시장 우위 추정"
    elif op_margin and op_margin >= 5:  pos, pos_txt = "Challenger", "안정적 수익 기반의 도전자 포지션"
    else:                               pos, pos_txt = "Niche",      "수익성 개선 또는 틈새 전략 필요"
    return {
        "industry": "정보 없음 (DART 사업보고서 확인 필요)",
        "market_dynamics": "공개 재무 데이터만으로는 산업 동학 분석 한계",
        "moat": "재무비율 기반 경쟁 해자 추정 — 상세 실사 필요",
        "growth_vectors": ["매출 성장률 기반 성장 동력 추정", "사업보고서 검토 필요", "업종별 드라이버 확인 권장"],
        "threats": ["시장 포화 가능성", "외부 경쟁 압력 증가"],
        "positioning": pos,
        "positioning_text": pos_txt,
    }


def _deal_memo_mock(corp_name: str, deal_type: str, yearly: dict) -> str:
    from datetime import date as _date
    today = _date.today().strftime("%Y년 %m월 %d일")
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default="N/A")
    d = yearly.get(latest_yr, {})

    def _f(v):
        if v is None: return "N/A"
        v2 = v / 1e8
        if abs(v2) >= 10000: return f"{v2/10000:.1f}조원"
        return f"{v2:.1f}억원"
    def _p(v): return f"{v:.1f}%" if v is not None else "N/A"

    return f"""# Deal Memo: {corp_name}

**딜 유형:** {deal_type} | **작성일:** {today} | **기밀 등급:** Confidential

---

## 1. Executive Summary

{corp_name}은(는) {latest_yr}년 기준 매출 {_f(d.get("매출"))}, 영업이익 {_f(d.get("영업이익"))}, 당기순이익 {_f(d.get("당기순이익"))}을 기록한 기업입니다. {deal_type} 관점의 본 메모는 DART 공시 재무 데이터 기반으로 작성된 예비 검토 자료이며, 본격적인 실사(DD) 이전 단계입니다.

> ⚠️ Claude CLI가 연결되지 않아 자동 생성된 Mock 메모입니다. Claude CLI 연동 시 AI가 전문적으로 작성합니다.

## 2. Why Now?

- {latest_yr}년 결산 데이터 공시 완료 — 초기 재무 검토 가능 시점
- {deal_type} 딜 구조 검토를 위한 기초 재무 데이터 확보
- 추가 실사(Due Diligence) 진행 전 예비 스크리닝 단계

## 3. Company Overview

- **기업명:** {corp_name}
- **분석 기준:** {latest_yr}년 연간 결산
- **총자산:** {_f(d.get("자산총계"))}
- **자본총계:** {_f(d.get("자본총계"))}
- **이익잉여금:** {_f(d.get("이익잉여금"))}

## 4. Industry View

공개 재무제표 데이터만으로는 세부 산업 분류가 어렵습니다. DART 사업보고서의 업종 코드 및 주요 사업 내용을 확인하여 정확한 산업 포지셔닝을 파악하세요.

## 5. Financial Checkpoints

| 지표 | {latest_yr}년 |
|------|--------------|
| 매출 | {_f(d.get("매출"))} |
| 영업이익 | {_f(d.get("영업이익"))} |
| 당기순이익 | {_f(d.get("당기순이익"))} |
| 영업이익률 | {_p(d.get("영업이익률"))} |
| ROE | {_p(d.get("ROE"))} |
| ROA | {_p(d.get("ROA"))} |
| 부채비율 | {_p(d.get("부채비율"))} |
| 유동비율 | {_p(d.get("유동비율"))} |

## 6. Deal Angle

{deal_type} 관점의 상세 딜 구조 분석은 Claude CLI 연동 후 AI 자동 생성을 권장합니다. 기초 재무 지표를 바탕으로 DCF/멀티플 기반 밸류에이션 작업을 우선 진행하세요.

## 7. Key Due Diligence Questions

1. 매출 구성 및 주요 고객사 현황 (매출 집중도 확인)
2. 미래 수주 잔고 및 파이프라인
3. 핵심 인력 구성 및 이탈 리스크
4. 기존 부채 구조, 만기 일정 및 리파이낸싱 계획
5. 법적 분쟁, 규제 리스크 및 지적재산권 현황
6. 관계사 거래 및 내부거래 현황
7. 환경/ESG 관련 잠재 비용 및 리스크

## 8. Risks

- **데이터 한계 (상):** 공시 재무제표 기반으로 관리 회계 정보 미포함
- **업종 불확실성 (중):** 세부 사업 내용 및 산업 동향 추가 확인 필요
- **시장 리스크 (중):** 금리, 환율 등 거시 경제 변수 영향 가능성
- **실사 리스크 (하):** 공시되지 않은 우발채무 또는 법적 이슈 존재 가능

## 9. Next Actions

1. DART 사업보고서 전문 다운로드 및 검토
2. 경영진 면담(Management Meeting) 일정 협의
3. 외부 법무/회계 자문사 선정 및 실사 범위 협의
4. 상세 재무 모델(DCF, 비교 멀티플) 작성
5. 초기 Term Sheet 초안 작성 및 내부 승인 요청

---
*본 메모는 DART 공시 재무 데이터 기반으로 자동 생성된 예비 자료입니다. 투자 결정 전 반드시 전문가 검토 및 실사를 진행하세요.*
"""


# ── Deal Intelligence: Claude Functions ───────────────────────────────────────
def generate_deal_signal_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = DEAL_SIGNAL_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    if raw.get("signal") not in {"Strong Buy", "Watch", "Caution", "Pass"}:
        raw["signal"] = "Watch"
    if raw.get("signal_level") not in {"good", "note", "warn", "danger"}:
        raw["signal_level"] = "note"
    if not isinstance(raw.get("reasons"), list):   raw["reasons"]    = []
    if not isinstance(raw.get("deal_types"), list): raw["deal_types"] = []
    return raw


def generate_industry_radar_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = INDUSTRY_RADAR_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    if raw.get("positioning") not in {"Leader", "Challenger", "Niche", "Vulnerable"}:
        raw["positioning"] = "Challenger"
    if not isinstance(raw.get("growth_vectors"), list): raw["growth_vectors"] = []
    if not isinstance(raw.get("threats"), list):        raw["threats"]        = []
    return raw


def generate_deal_memo_claude(corp_name: str, deal_type: str, yearly: dict) -> str | None:
    if not check_claude_cli(): return None
    from datetime import date as _date
    prompt = DEAL_MEMO_PROMPT.format(
        fin_table=_build_fin_table(corp_name, yearly),
        corp_name=corp_name, deal_type=deal_type,
        date=_date.today().strftime("%Y년 %m월 %d일"),
    )
    response = call_claude_cli(prompt, timeout=180)
    if not response: return None
    text = response.strip()
    if text.startswith("```markdown"): text = text[11:]
    elif text.startswith("```"):       text = text[3:]
    if text.endswith("```"):           text = text[:-3]
    return text.strip()




# ── 산업 인텔리전스: Naver News API ───────────────────────────────────────────
NAVER_CLIENT_ID     = os.getenv("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET", "")
NEWS_API_KEY        = os.getenv("NEWS_API_KEY", "")  # newsapi.org

# ── 산업 키워드 한→영 매핑 ────────────────────────────────────────────────────
INDUSTRY_KW_MAP: dict[str, list[str]] = {
    "반도체":       ["semiconductor investment", "chip industry M&A", "AI chip capex", "HBM market", "memory chip outlook"],
    "AI":          ["artificial intelligence investment", "LLM market", "generative AI", "AI infrastructure M&A", "AI data center"],
    "2차전지":      ["battery market", "EV battery demand", "lithium price", "battery materials", "cathode anode supply chain"],
    "전기차":       ["electric vehicle market", "EV demand outlook", "automotive electrification", "EV charging infrastructure"],
    "핀테크":       ["fintech investment", "digital banking", "payment technology M&A", "neobank", "embedded finance"],
    "클라우드":     ["cloud computing market", "SaaS M&A", "cloud infrastructure", "hyperscaler capex"],
    "SaaS":        ["SaaS M&A", "B2B software market", "ARR growth", "software valuation", "cloud software"],
    "바이오":       ["biotechnology M&A", "biopharma deal", "drug development", "biotech investment", "clinical trial"],
    "바이오텍":     ["biotech M&A", "biopharma investment", "FDA approval", "drug licensing deal"],
    "의료기기":     ["medical device M&A", "medtech investment", "digital health", "FDA clearance", "hospital spending"],
    "디지털헬스":   ["digital health investment", "healthtech M&A", "telehealth market", "wearable health"],
    "게임":         ["gaming industry M&A", "mobile game market", "esports investment", "game studio acquisition"],
    "태양광":       ["solar energy investment", "photovoltaic market", "renewable energy capex", "solar panel supply"],
    "수소":         ["hydrogen energy investment", "green hydrogen", "fuel cell market", "hydrogen economy"],
    "2차전지소재":  ["battery materials M&A", "lithium supply chain", "cathode material", "battery recycling"],
    "철강":         ["steel industry outlook", "steel M&A", "iron ore price", "steel capex"],
    "석유화학":     ["petrochemical market", "chemical industry M&A", "polymer demand", "ethylene spread"],
    "디스플레이":   ["display panel market", "OLED investment", "display M&A", "panel supply chain"],
    "자동차":       ["automotive M&A", "car industry outlook", "auto supplier", "autonomous driving investment"],
    "로봇":         ["robotics investment", "automation M&A", "industrial robot market", "humanoid robot"],
    "항공우주":     ["aerospace M&A", "defense investment", "aviation industry", "space tech"],
    "방산":         ["defense M&A", "defense spending", "military technology", "defense contract"],
    "패션":         ["fashion M&A", "retail investment", "luxury brand acquisition", "fashion tech"],
    "뷰티":         ["beauty industry M&A", "cosmetics investment", "K-beauty global", "beauty brand acquisition"],
    "e커머스":      ["e-commerce M&A", "online retail investment", "marketplace acquisition", "logistics tech"],
    "유통":         ["retail M&A", "distribution investment", "logistics acquisition", "supply chain investment"],
    "엔터테인먼트": ["entertainment M&A", "media investment", "content acquisition", "streaming deal"],
    "부동산":       ["real estate M&A", "proptech investment", "REIT market", "commercial real estate"],
    "은행":         ["banking M&A", "financial services deal", "bank acquisition", "fintech banking"],
    "보험":         ["insurance M&A", "insurtech investment", "reinsurance deal"],
    "모빌리티":     ["mobility investment", "autonomous driving M&A", "ride-hailing market", "fleet management"],
    "스마트팩토리": ["smart factory investment", "industrial IoT", "factory automation M&A", "manufacturing tech"],
    "CRO":         ["CRO M&A", "contract research organization", "CDMO deal", "pharma outsourcing"],
    "CDMO":        ["CDMO investment", "contract manufacturing M&A", "biomanufacturing", "pharma CDMO"],
    "화장품·식품":  ["cosmetics M&A", "K-beauty global", "food beverage M&A", "consumer staples investment", "FMCG outlook"],
    "건설기계":     ["construction equipment market", "heavy machinery M&A", "earthmoving equipment", "Caterpillar Komatsu outlook"],
}



# ── 기업명 영문 변환 ───────────────────────────────────────────────────────────
CORP_NAME_EN_MAP: dict[str, str] = {
    # 국내 대기업
    "삼성전자": "Samsung Electronics", "삼성": "Samsung",
    "SK하이닉스": "SK Hynix", "SK텔레콤": "SK Telecom", "SK이노베이션": "SK Innovation",
    "SK바이오사이언스": "SK Bioscience", "SK": "SK Group",
    "LG전자": "LG Electronics", "LG화학": "LG Chem",
    "LG에너지솔루션": "LG Energy Solution", "LG디스플레이": "LG Display", "LG": "LG",
    "현대자동차": "Hyundai Motor", "현대차": "Hyundai Motor",
    "기아": "Kia", "현대모비스": "Hyundai Mobis", "현대": "Hyundai",
    "카카오": "Kakao", "카카오뱅크": "Kakao Bank", "카카오페이": "Kakao Pay",
    "카카오모빌리티": "Kakao Mobility", "카카오게임즈": "Kakao Games",
    "네이버": "NAVER", "네이버클라우드": "NAVER Cloud", "네이버웹툰": "NAVER Webtoon",
    "셀트리온": "Celltrion", "삼성바이오로직스": "Samsung Biologics",
    "삼성SDI": "Samsung SDI", "삼성물산": "Samsung C&T",
    "포스코": "POSCO", "포스코홀딩스": "POSCO Holdings", "포스코퓨처엠": "POSCO Future M",
    "한국전력": "KEPCO", "KT": "KT", "KT&G": "KT&G",
    "크래프톤": "Krafton", "하이브": "HYBE", "엔씨소프트": "NCSOFT",
    "넥슨": "Nexon", "넷마블": "Netmarble", "펄어비스": "Pearl Abyss",
    "에코프로비엠": "EcoPro BM", "에코프로": "EcoPro",
    "엘앤에프": "L&F", "포스코케미칼": "POSCO Chemical",
    "두산에너빌리티": "Doosan Enerbility", "두산로보틱스": "Doosan Robotics",
    "한화솔루션": "Hanwha Solutions", "한화에어로스페이스": "Hanwha Aerospace",
    "한화오션": "Hanwha Ocean", "한화": "Hanwha",
    "롯데케미칼": "Lotte Chemical", "롯데": "Lotte",
    "CJ제일제당": "CJ CheilJedang", "CJ대한통운": "CJ Logistics", "CJ": "CJ Group",
    "신한금융": "Shinhan Financial", "KB금융": "KB Financial",
    "하나금융": "Hana Financial", "우리금융": "Woori Financial",
    "토스": "Toss", "토스뱅크": "Toss Bank",
    "쿠팡": "Coupang", "컬리": "Kurly", "오아시스": "Oasis",
    "배달의민족": "Baemin", "야놀자": "Yanolja",
    "카카카오": "Kakao",  # 오타 방어
    # 외국 기업 한글 표기
    "엔비디아": "NVIDIA", "애플": "Apple", "구글": "Google",
    "알파벳": "Alphabet", "마이크로소프트": "Microsoft", "MS": "Microsoft",
    "아마존": "Amazon", "AWS": "AWS", "메타": "Meta",
    "테슬라": "Tesla", "인텔": "Intel", "퀄컴": "Qualcomm",
    "브로드컴": "Broadcom", "ARM": "ARM", "TSMC": "TSMC",
    "소프트뱅크": "SoftBank", "알리바바": "Alibaba",
    "텐센트": "Tencent", "바이두": "Baidu", "화웨이": "Huawei",
    "삼성전자우": "Samsung Electronics",
}

def _is_korean(text: str) -> bool:
    """한글 포함 여부 확인."""
    return any('\uAC00' <= c <= '\uD7A3' or '\u1100' <= c <= '\u11FF' for c in text)

def _corp_to_en(corp_name: str) -> str:
    """한글 기업명을 영문으로 변환. 이미 영문이면 그대로 반환."""
    name = corp_name.strip()
    if not _is_korean(name):
        return name  # 이미 영문
    return CORP_NAME_EN_MAP.get(name, "")  # 매핑 없으면 빈 문자열

def _build_global_kws(corp_name: str, industry: str) -> list[str]:
    """글로벌 뉴스 검색용 영문 키워드 목록 구성."""
    kws: list[str] = []
    industry_kws = INDUSTRY_KW_MAP.get((industry or "").strip(), [])

    en_corp = _corp_to_en(corp_name) if corp_name else ""

    if en_corp:
        kws.append(en_corp)                          # 기업명 단독 (가장 중요)
        kws.append(f"{en_corp} investment")
        kws.append(f"{en_corp} M&A")
        if industry_kws:
            kws.append(f"{en_corp} {industry_kws[0].split()[0]}")  # e.g. "Kakao fintech"
    elif not _is_korean(corp_name or ""):
        # 이미 영문 기업명 (NVIDIA 등)
        kws.append(corp_name.strip())
        kws.append(f"{corp_name.strip()} investment")

    # 산업 키워드 추가 (중복 제외)
    for kw in industry_kws[:3]:
        if kw not in kws:
            kws.append(kw)

    # 키워드가 없으면 산업명 기반 fallback
    if not kws and industry:
        kws = _get_en_keywords(industry)

    return kws[:5]

def _get_en_keywords(industry: str) -> list[str]:
    """산업명에서 영어 검색 키워드 반환."""
    kws = INDUSTRY_KW_MAP.get((industry or "").strip())
    if kws:
        return kws[:4]
    en = (industry or "").strip()
    if not en or _is_korean(en):
        return []   # 한글 산업명 매핑 없으면 빈 리스트 반환
    return [f"{en} investment", f"{en} M&A", f"{en} market", f"{en} industry"]


def _extract_domain(url: str) -> str:
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc
        return host.replace("www.", "").split(".")[0] if host else ""
    except Exception:
        return ""


def search_naver_news(query: str, display: int = 15) -> list:
    """Naver News Search API. API 키 없으면 빈 리스트."""
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        return []
    try:
        r = requests.get(
            "https://openapi.naver.com/v1/search/news.json",
            headers={
                "X-Naver-Client-Id":     NAVER_CLIENT_ID,
                "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
            },
            params={"query": query, "display": display, "sort": "date"},
            timeout=10,
        )
        if r.status_code != 200:
            return []
        return [
            {
                "title":           re.sub(r'<[^>]+>', '', item.get("title", "")),
                "description":     re.sub(r'<[^>]+>', '', item.get("description", "")),
                "url":             item.get("originallink") or item.get("link", ""),
                "source":          _extract_domain(item.get("originallink") or item.get("link", "")),
                "published_at":    item.get("pubDate", "")[:16],
                "source_language": "ko",
            }
            for item in r.json().get("items", [])
        ]
    except Exception:
        return []


def search_newsapi(query: str, display: int = 10) -> list:
    """NewsAPI (newsapi.org). NEWS_API_KEY 없으면 빈 리스트."""
    if not NEWS_API_KEY:
        return []
    try:
        r = requests.get(
            "https://newsapi.org/v2/everything",
            params={
                "q": query, "language": "en", "sortBy": "publishedAt",
                "pageSize": min(display, 20), "apiKey": NEWS_API_KEY,
            },
            timeout=10,
        )
        if r.status_code != 200:
            return []
        return [
            {
                "title":        a.get("title", ""),
                "description":  a.get("description") or "",
                "url":          a.get("url", ""),
                "source":       a.get("source", {}).get("name", ""),
                "published_at": (a.get("publishedAt") or "")[:10],
                "source_language": "en",
            }
            for a in r.json().get("articles", [])
            if a.get("title") and "[Removed]" not in a.get("title", "")
        ]
    except Exception:
        return []


def search_gdelt(query: str, display: int = 10) -> list:
    """GDELT DOC API 2.0 — 별도 API 키 불필요."""
    try:
        r = requests.get(
            "https://api.gdeltproject.org/api/v2/doc/doc",
            params={
                "query": query, "mode": "artlist",
                "maxrecords": min(display, 25), "format": "json",
                "sort": "DateDesc",
            },
            timeout=12,
        )
        if r.status_code != 200:
            return []
        articles = r.json().get("articles") or []
        return [
            {
                "title":        a.get("title", ""),
                "description":  "",
                "url":          a.get("url", ""),
                "source":       _extract_domain(a.get("url", "")),
                "published_at": (a.get("seendate") or "")[:10],
                "source_language": "en",
            }
            for a in articles
            if a.get("title") and a.get("url")
        ]
    except Exception:
        return []


def search_global_news(en_keywords: list[str], display_per_kw: int = 5) -> list:
    """NewsAPI 우선, 없으면 GDELT fallback. 병렬 수집."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    kws = en_keywords[:3]
    seen_urls: set = set()
    results: list = []
    def _fetch(kw):
        arts = search_newsapi(kw, display_per_kw)
        if not arts:
            arts = search_gdelt(kw, display_per_kw)
        return arts
    with ThreadPoolExecutor(max_workers=3) as ex:
        futures = {ex.submit(_fetch, kw): kw for kw in kws}
        for fut in as_completed(futures, timeout=15):
            try:
                for a in (fut.result() or []):
                    if a.get("url") and a["url"] not in seen_urls:
                        seen_urls.add(a["url"])
                        results.append(a)
            except Exception:
                pass
    return results[:display_per_kw * 3]


def _fmt_news_for_claude(news: list, text_input: str = "") -> str:
    """뉴스 목록을 Claude 프롬프트용 텍스트로 변환. 언어 태그 포함."""
    ko_parts, en_parts = [], []
    for i, n in enumerate(news[:20], 1):
        lang = n.get("source_language", "ko")
        line = f"{i}. [{n.get('source','미상')}] {n['title']}"
        if n.get("description"):
            line += f"\n   {n['description'][:150]}"
        if n.get("published_at"):
            line += f" ({n['published_at']})"
        if lang == "en":
            en_parts.append(line)
        else:
            ko_parts.append(line)
    parts = []
    if ko_parts:
        parts.append("[국내 뉴스]")
        parts.extend(ko_parts)
    if en_parts:
        parts.append("\n[글로벌/영문 뉴스]")
        parts.extend(en_parts)
    if text_input and text_input.strip():
        parts.append("\n[사용자 입력 본문]")
        parts.append(text_input.strip()[:4000])
    return "\n".join(parts) if parts else "뉴스/본문 데이터 없음"


# ── 산업 인텔리전스: Claude 프롬프트 ──────────────────────────────────────────
COMPANY_TRENDS_PROMPT = """\
당신은 VC/M&A/PE 전문 산업 분석 어드바이저입니다.
아래 뉴스/자료와 재무 데이터를 기반으로 {corp_name}이 속한 산업 동향과 딜 인사이트를 분석하세요.
JSON으로만 출력하세요. JSON 이외 텍스트는 절대 출력하지 마세요.

[뉴스/자료]
{news_text}

[재무 데이터]
{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "core_signals":           ["핵심 산업 시그널 1 (25자)", "시그널 2", "시그널 3"],
  "recent_issues":          [{{"title":"이슈명(20자)","summary":"30자 요약"}}, ...],
  "impact_on_corp":         "{corp_name}에 미치는 영향 (60자)",
  "deal_angle":             "딜 관점 코멘트 VC/M&A/PE 중 해당 관점 (50자)",
  "financial_impact":       "재무지표 변화 예상 (40자)",
  "accounting_checkpoints": ["회계·공시 체크포인트 1 (25자)", "2", "3"],
  "dd_questions":           ["실사 필수 질문 1 (30자)", "2", "3"],
  "risks":                  [{{"factor":"리스크명(20자)","level":"상|중|하"}}]
}}"""

INDUSTRY_ANALYSIS_PROMPT = """\
당신은 투자은행 시니어 산업 리서치 애널리스트입니다.
아래 뉴스/자료를 기반으로 {industry} 산업을 종합 분석하세요.
JSON으로만 출력하세요. JSON 이외 텍스트는 절대 출력하지 마세요.

[뉴스/자료]
{news_text}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "sector_overview":    "섹터 개요 한 문장 (70자)",
  "recent_trends":      ["최근 트렌드 1 (30자)", "2", "3"],
  "growth_drivers":     ["성장 동력 1 (30자)", "2", "3"],
  "risks":              ["주요 리스크 1 (30자)", "2", "3"],
  "major_players":      ["주요 플레이어/세그먼트 1", "2", "3"],
  "deal_opportunities": ["딜 기회 1 (40자)", "딜 기회 2"],
  "valuation_points":   ["밸류에이션 포인트 1 (35자)", "포인트 2"],
  "accounting_points":  ["회계 포인트 1 (35자)", "포인트 2"],
  "dd_questions":       ["실사 질문 1 (35자)", "질문 2", "질문 3"]
}}"""


# ── 산업 인텔리전스: Mock ──────────────────────────────────────────────────────
def _company_trends_mock(corp_name: str, industry: str, has_data: bool) -> dict:
    hint = "" if has_data else " (뉴스 API 키 또는 본문 입력 필요)"
    return {
        "domestic_issues":        [{"title": "국내 데이터 없음", "summary": f"Naver API 키 설정 시 자동 수집{hint}"}],
        "global_issues":          [{"title": "글로벌 데이터 없음", "summary": "NEWS_API_KEY 또는 GDELT로 수집 가능"}],
        "common_signals":         [
            f"{industry or corp_name + ' 관련 산업'} 동향 분석 준비 중{hint}",
            "NAVER_CLIENT_ID/SECRET + NEWS_API_KEY .env 설정 권장",
            "Claude CLI 연동 시 AI 딜 분석 자동 생성",
        ],
        "impact_on_corp":         f"{corp_name}에 대한 분석은 뉴스 데이터 수집 후 제공됩니다",
        "deal_angle":             "뉴스 수집 후 VC/M&A/PE 관점 딜 각도 분석 제공",
        "financial_impact":       "데이터 입력 후 재무 영향 자동 분석",
        "accounting_checkpoints": ["수익 인식 정책 검토", "주요 자산 손상 여부 확인", "우발채무 공시 검토"],
        "dd_questions":           ["핵심 사업 모델 및 경쟁 우위는?", "최근 3년 매출 구성 변화 원인은?", "주요 고객사 집중도 및 계약 현황은?"],
        "risks":                  [{"factor": "데이터 부족", "level": "중"}, {"factor": "시장 불확실성", "level": "중"}],
    }


def _industry_analysis_mock(industry: str, has_data: bool) -> dict:
    hint = "" if has_data else " (뉴스 API 키 또는 본문 입력 필요)"
    return {
        "sector_overview":    f"{industry} 산업 분석을 위해 뉴스 데이터 입력이 필요합니다{hint}",
        "domestic_issues":    [{"title": "국내 데이터 없음", "summary": f"Naver API 키 설정 시 자동 수집{hint}"}],
        "global_issues":      [{"title": "글로벌 데이터 없음", "summary": "NEWS_API_KEY .env 추가 또는 GDELT 자동 수집"}],
        "common_signals":     ["뉴스 수집 후 국내외 공통 시그널 분석", "API 키 설정 또는 본문 직접 입력 권장"],
        "deal_angle":         "뉴스 수집 후 딜 각도 분석 제공",
        "deal_opportunities": ["뉴스 분석 후 제공"],
        "valuation_points":   ["데이터 입력 후 분석"],
        "accounting_points":  ["데이터 입력 후 분석"],
        "dd_questions":       ["산업 내 경쟁 구도 변화는?", "주요 성장 드라이버 지속 가능성은?", "규제 환경 변화 리스크는?"],
    }


# ── 산업 인텔리전스: Claude 생성 함수 ─────────────────────────────────────────
def generate_company_trends_claude(corp_name: str, news: list, text_input: str, fin_table: str) -> dict | None:
    if not check_claude_cli(): return None
    news_text = _fmt_news_for_claude(news, text_input)
    if news_text == "뉴스/본문 데이터 없음": return None
    prompt = COMPANY_TRENDS_PROMPT.format(
        corp_name=corp_name, news_text=news_text,
        fin_table=fin_table or "재무 데이터 없음",
    )
    response = call_claude_cli(prompt, timeout=120)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    for k in ["domestic_issues", "global_issues", "common_signals", "accounting_checkpoints", "dd_questions", "risks"]:
        if not isinstance(raw.get(k), list): raw[k] = []
    return raw


def generate_industry_analysis_claude(industry: str, news: list, text_input: str) -> dict | None:
    if not check_claude_cli(): return None
    news_text = _fmt_news_for_claude(news, text_input)
    if news_text == "뉴스/본문 데이터 없음": return None
    prompt = INDUSTRY_ANALYSIS_PROMPT.format(industry=industry, news_text=news_text)
    response = call_claude_cli(prompt, timeout=120)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    for k in ["domestic_issues","global_issues","common_signals","deal_opportunities","valuation_points","accounting_points","dd_questions"]:
        if not isinstance(raw.get(k), list): raw[k] = []
    return raw


# ── Excel Generation ───────────────────────────────────────────────────────────
def _xl_styles():
    def mf(bold=False, color="191F28", size=11, italic=False):
        return Font(name="맑은 고딕", bold=bold, color=color, size=size, italic=italic)
    def fl(hex_color):
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))
    BD   = Border(
        left=Side(style="thin", color="E5E8EB"),
        right=Side(style="thin", color="E5E8EB"),
        top=Side(style="thin", color="E5E8EB"),
        bottom=Side(style="thin", color="E5E8EB"),
    )
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    return mf, fl, BD, CTR, LEFT


def _write_company_sheet(ws, ci, corp_name, yearly_ci):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    co_hex = CO_COLS[ci]
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    ws.merge_cells("A1:G2")
    cell(1, 1, f"{corp_name}  재무분석", mf(True,"FFFFFF",14), co_hex, CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells("A3:G3")
    cell(3, 1, f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    for sec_title, keys, is_ratio in SECTIONS:
        ws.merge_cells(f"A{r}:G{r}")
        cell(r, 1, sec_title, mf(True,"FFFFFF",12), co_hex, CTR)
        ws.row_dimensions[r].height = 26; r += 1
        for ci2, h in enumerate(["항목","2023","2024","YoY","2025","YoY",""], 1):
            cell(r, ci2, h, mf(True,"FFFFFF",10), "6B7684", CTR)
        ws.row_dimensions[r].height = 20; r += 1
        for ki, key in enumerate(keys):
            vals = [yearly_ci.get(yr, {}).get(key)
                    if yearly_ci.get(yr) else None for yr in YEARS]
            bg = "F9FAFB" if ki % 2 == 0 else "FFFFFF"
            ws.row_dimensions[r].height = 22
            cell(r, 1, key, mf(color="6B7684"), bg, LEFT)
            cell(r, 2, fmt_val(vals[0], key, is_ratio), mf(True), bg, CTR)
            cell(r, 7, "", mf(), bg, CTR)
            for col_off, idx in [(3, 1), (5, 2)]:
                v   = vals[idx]
                txt = yoy_str(v, vals[idx-1])
                cell(r, col_off, fmt_val(v, key, is_ratio), mf(True), bg, CTR)
                yf  = (mf(color="05C072",size=10) if txt.startswith("▲")
                       else mf(color="F04452",size=10) if txt.startswith("▼")
                       else mf(color="B0B8C1",size=10))
                cell(r, col_off+1, txt or "-", yf, bg, CTR)
            r += 1
        r += 1
    for col_l, w in [("A",14),("B",14),("C",14),("D",11),("E",14),("F",11),("G",6)]:
        ws.column_dimensions[col_l].width = w


def _write_comparison_sheet(ws, cos, yearly):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    active   = [(ci, c["corp_name"]) for ci, c in enumerate(cos) if c]
    n        = len(active)
    last_col = 1 + n * 3 + 1
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    ws.merge_cells(f"A1:{chr(64+last_col)}2")
    cell(1, 1, "3사 재무 비교 분석", mf(True,"FFFFFF",14), "191F28", CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A3:{chr(64+last_col)}3")
    cell(3, 1, f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    for sec_title, keys, is_ratio in SECTIONS:
        ws.merge_cells(f"A{r}:{chr(64+last_col)}{r}")
        cell(r, 1, sec_title, mf(True,"FFFFFF",12), "4B9FFF", CTR)
        ws.row_dimensions[r].height = 26; r += 1
        cell(r, 1, "항목", mf(True,"FFFFFF",10), "6B7684", LEFT)
        for li, (ci, nm) in enumerate(active):
            base_c = 2 + li * 3
            ws.merge_cells(f"{chr(64+base_c)}{r}:{chr(64+base_c+2)}{r}")
            cell(r, base_c, nm, mf(True,"FFFFFF",11), CO_COLS[ci], CTR)
        ws.row_dimensions[r].height = 26; r += 1
        cell(r, 1, "", mf(True,"FFFFFF",9), "6B7684", CTR)
        for li, (ci, _) in enumerate(active):
            for yi, yr in enumerate(YEARS):
                cell(r, 2+li*3+yi, yr, mf(True,"FFFFFF",9), "6B7684", CTR)
        ws.row_dimensions[r].height = 18; r += 1
        for ki, key in enumerate(keys):
            bg = "F9FAFB" if ki % 2 == 0 else "FFFFFF"
            ws.row_dimensions[r].height = 22
            cell(r, 1, key, mf(color="6B7684"), bg, LEFT)
            for li, (ci, _) in enumerate(active):
                for yi, yr in enumerate(YEARS):
                    v = yearly[ci].get(yr, {}).get(key) if yearly[ci].get(yr) else None
                    cell(r, 2+li*3+yi, fmt_val(v, key, is_ratio), mf(True), bg, CTR)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 14
    for li in range(n):
        for yi in range(3):
            ws.column_dimensions[chr(66+li*3+yi)].width = 13


def _write_opinion_sheet(ws, cos, opinions):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    active = [(ci, c["corp_name"]) for ci, c in enumerate(cos) if c]
    n = len(active)
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    last_c = 2 + n
    ws.merge_cells(f"A1:{chr(64+last_c)}2")
    cell(1, 1, "재무 분석 의견 — Claude AI 자동 평가",
         mf(True,"FFFFFF",14), "191F28", CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A3:{chr(64+last_c)}3")
    cell(3, 1, f"생성일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  Claude CLI",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    cell(r, 1, "분류",    mf(True,"FFFFFF",10), "6B7684", CTR)
    cell(r, 2, "평가 항목", mf(True,"FFFFFF",10), "6B7684", CTR)
    for li, (ci, nm) in enumerate(active):
        cell(r, 3+li, nm, mf(True,"FFFFFF",11), CO_COLS[ci], CTR)
    ws.row_dimensions[r].height = 22; r += 1
    prev_cat = None
    for cat, lbl in OPINION_KEYS:
        bg_lbl = "EFF6FF" if cat != prev_cat else "FFFFFF"
        prev_cat = cat
        ws.row_dimensions[r].height = 54
        cell(r, 1, cat, mf(True,"6B7684",10), bg_lbl, CTR)
        cell(r, 2, lbl, mf(bold=True,size=11), "FFFFFF", LEFT)
        for li, (ci, _) in enumerate(active):
            op   = opinions[ci] or {}
            item = op.get(lbl, {"value":"N/A","text":"데이터 없음","level":"note"})
            lv   = item.get("level","note")
            combined = f"{LV_ICON.get(lv,'·')} {item.get('value','N/A')}\n{item.get('text','')}"
            cc   = ws.cell(r, 3+li, combined)
            cc.font      = Font(name="맑은 고딕", size=10, color=LV_XL.get(lv,"B0B8C1"),
                                bold=(lv=="good"))
            cc.fill      = PatternFill("solid", fgColor=LV_XL_BG.get(lv,"FFFFFF"))
            cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cc.border    = BD
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for li in range(n):
        ws.column_dimensions[chr(67+li)].width = 44


def generate_excel(cos: list, yearly: list, opinions: list | None = None) -> str:
    if opinions is None:
        opinions = [None, None, None]
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    names    = [c["corp_name"] for c in cos if c]
    stem     = "_".join(names[:2]) + ("_외" if len(names) > 2 else "")
    filename = f"{stem}_재무비교_{ts}.xlsx"
    path     = os.path.join(tempfile.gettempdir(), filename)
    wb       = openpyxl.Workbook(); wb.remove(wb.active)
    for ci, corp in enumerate(cos):
        if not corp: continue
        ws = wb.create_sheet(title=corp["corp_name"][:28])
        _write_company_sheet(ws, ci, corp["corp_name"], yearly[ci])
    ws_cmp = wb.create_sheet(title="전체 비교")
    _write_comparison_sheet(ws_cmp, cos, yearly)
    if any(opinions):
        ws_op = wb.create_sheet(title="재무분석 의견")
        _write_opinion_sheet(ws_op, cos, opinions)
    for target in reversed(["재무분석 의견", "전체 비교"]):
        if target in wb.sheetnames:
            wb.move_sheet(target, offset=-len(wb.sheetnames))
    wb.save(path)
    return filename


# ── FastAPI App ────────────────────────────────────────────────────────────────
app = FastAPI(title="DART 재무분석기 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_dart      = DART()
_dart_lock = threading.Lock()
_dart_ready = False
_pool      = concurrent.futures.ThreadPoolExecutor(max_workers=6)


def _ensure_dart():
    if not API_KEY:
        raise HTTPException(status_code=500, detail="DART_API_KEY가 설정되지 않았습니다. .env 파일을 확인하세요.")
    global _dart_ready
    if not _dart_ready:
        with _dart_lock:
            if not _dart_ready:
                _dart.load()
                _dart_ready = True


# ── Pydantic Models ────────────────────────────────────────────────────────────
class SearchReq(BaseModel):
    query: str

class CorpInfo(BaseModel):
    corp_name:  str
    corp_code:  str
    stock_code: str = ""

class AnalyzeReq(BaseModel):
    corps: list[Optional[CorpInfo]]

class OpinionReq(BaseModel):
    corp_name: str
    yearly:    dict


# ── Endpoints ──────────────────────────────────────────────────────────────────
@app.post("/api/search")
async def search_corp(body: SearchReq):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(_pool, _ensure_dart)
    return {"results": _dart.search(body.query)}


@app.post("/api/analyze-multi")
async def analyze_multi(body: AnalyzeReq):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(_pool, _ensure_dart)

    def _run():
        cos    = [c.model_dump() if c else None for c in body.corps]
        yearly = [{}, {}, {}]
        dart_status = [{"found": False, "has_financials": False} for _ in range(3)]
        for ci, corp in enumerate(cos):
            if not corp: continue
            cc = corp.get("corp_code", "")
            # 비상장 직접 입력인 경우 DART 조회 생략
            if not cc or cc == "__unlisted__":
                dart_status[ci] = {"found": False, "has_financials": False}
                continue
            dart_status[ci]["found"] = True
            for yr in YEARS:
                items, fs = _dart.financials(cc, yr)
                if items:
                    fin = _dart.extract(items)
                    rat = DART.calc_ratios(fin)
                    yearly[ci][yr] = {**fin, **rat, "_fs": fs}
            if any(yearly[ci].values()):
                dart_status[ci]["has_financials"] = True
        filename = generate_excel(cos, yearly)
        return {"yearly": yearly, "filename": filename, "dart_status": dart_status}

    return await loop.run_in_executor(_pool, _run)


@app.post("/api/opinion")
async def get_opinion(body: OpinionReq):
    if not check_claude_cli():
        raise HTTPException(status_code=503, detail="Claude CLI 미설치")
    loop   = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _pool, lambda: generate_opinions_claude(body.corp_name, body.yearly)
    )
    if result is None:
        raise HTTPException(status_code=504, detail="Claude CLI 응답 없음")
    return result


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(tempfile.gettempdir(), safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")
    return FileResponse(
        path, filename=safe,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




# ── Deal Intelligence Models ───────────────────────────────────────────────────
class DealSignalReq(BaseModel):
    corp_name: str
    yearly:    dict

class IndustryRadarReq(BaseModel):
    corp_name: str
    yearly:    dict

class DealMemoReq(BaseModel):
    corp_name: str
    deal_type: str = "M&A"
    yearly:    dict


# ── Deal Intelligence Endpoints ────────────────────────────────────────────────
@app.post("/api/deal-signal")
async def deal_signal(body: DealSignalReq):
    loop = asyncio.get_event_loop()
    def _run():
        result = generate_deal_signal_claude(body.corp_name, body.yearly)
        return result or _deal_signal_mock(body.corp_name, body.yearly)
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/industry-radar")
async def industry_radar(body: IndustryRadarReq):
    loop = asyncio.get_event_loop()
    def _run():
        result = generate_industry_radar_claude(body.corp_name, body.yearly)
        return result or _industry_radar_mock(body.corp_name, body.yearly)
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/deal-memo")
async def deal_memo_endpoint(body: DealMemoReq):
    loop = asyncio.get_event_loop()
    def _run():
        memo_text = generate_deal_memo_claude(body.corp_name, body.deal_type, body.yearly)
        if not memo_text:
            memo_text = _deal_memo_mock(body.corp_name, body.deal_type, body.yearly)
        memo_dir = os.path.join(tempfile.gettempdir(), "memos")
        os.makedirs(memo_dir, exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_nm  = re.sub(r'[^\w가-힣-]', '_', body.corp_name)
        filename = f"{safe_nm}_{body.deal_type}_{ts}.md"
        with open(os.path.join(memo_dir, filename), "w", encoding="utf-8") as fh:
            fh.write(memo_text)
        return {"memo": memo_text, "filename": filename}
    return await loop.run_in_executor(_pool, _run)


@app.get("/api/download-memo/{filename}")
async def download_memo_file(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(tempfile.gettempdir(), "memos", safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="메모 파일을 찾을 수 없습니다")
    return FileResponse(path, filename=safe, media_type="text/markdown; charset=utf-8")





# ── 비상장/비공시 기업 수동 분석 ─────────────────────────────────────────────
class ManualFinReq(BaseModel):
    corp_name:  str
    industry:   str  = ""
    deal_type:  str  = "M&A"
    year:       str  = ""
    revenue:              Optional[float] = None
    operating_profit:     Optional[float] = None
    net_income:           Optional[float] = None
    total_assets:         Optional[float] = None
    total_liabilities:    Optional[float] = None
    equity:               Optional[float] = None
    current_assets:       Optional[float] = None
    current_liabilities:  Optional[float] = None
    notes:      str  = ""


MANUAL_DEAL_MEMO_PROMPT = """당신은 M&A·VC·PE 딜 자문 전문가입니다.
아래 기업 정보와 수동 입력 재무 데이터를 바탕으로 Deal Memo를 작성하세요.

기업명: {corp_name}
산업/섹터: {industry}
딜 유형: {deal_type}
기준 연도: {year}

재무 데이터 (단위: 백만원):
{fin_table}

재무비율:
{ratio_table}

추가 메모: {notes}

아래 구조로 한국어 Deal Memo를 작성하세요. 마크다운 형식 사용:
# Deal Memo — {corp_name} ({deal_type})

## 1. 기업 개요
(산업, 비즈니스 모델, 포지셔닝 요약)

## 2. 재무 하이라이트
(입력된 재무 데이터 기반 핵심 수치 요약)

## 3. 재무 건전성 평가
(제공된 재무비율 기반 수익성·안정성·성장성 평가)

## 4. 투자/딜 관점 ({deal_type})
(딜 구조, 가치평가 시사점, 리스크)

## 5. 핵심 리스크
(비상장·비공시 기업 특유 리스크 포함)

## 6. DD 체크리스트
(최소 5개 핵심 DD 항목)

## 7. 결론 및 권고
(Go/Caution/Pass 판단 + 근거)

⚠️ 본 메모는 제한된 공개 데이터 기반 작성되었으며, 투자 권유가 아닙니다.
"""


def _calc_ratios_manual(req: ManualFinReq) -> dict:
    rev  = req.revenue        or 0
    op   = req.operating_profit
    net  = req.net_income
    ast  = req.total_assets   or 0
    eq   = req.equity         or 0
    dbt  = req.total_liabilities or 0
    ca   = req.current_assets or 0
    cl   = req.current_liabilities or 0
    def pct(a, b): return round(a / b * 100, 2) if b else None
    def rat(a, b): return round(a / b, 2) if b else None
    return {
        "영업이익률":  pct(op, rev)  if op is not None else None,
        "순이익률":    pct(net, rev) if net is not None else None,
        "ROE":        pct(net, eq)  if net is not None else None,
        "ROA":        pct(net, ast) if net is not None else None,
        "부채비율":    pct(dbt, eq) if eq else None,
        "유동비율":    pct(ca, cl)  if cl else None,
    }


def _fmt_manual_fin(req: ManualFinReq) -> tuple[str, str]:
    fields = [
        ("매출", req.revenue),
        ("영업이익", req.operating_profit),
        ("당기순이익", req.net_income),
        ("자산총계", req.total_assets),
        ("부채총계", req.total_liabilities),
        ("자본총계", req.equity),
        ("유동자산", req.current_assets),
        ("유동부채", req.current_liabilities),
    ]
    fin_lines = [f"- {k}: {int(v):,}" for k, v in fields if v is not None]
    fin_table = "\n".join(fin_lines) if fin_lines else "재무 데이터 없음"
    ratios = _calc_ratios_manual(req)
    ratio_lines = [f"- {k}: {v:.2f}{'%' if '율' in k or 'ROE' == k or 'ROA' == k else 'x'}"
                   for k, v in ratios.items() if v is not None]
    ratio_table = "\n".join(ratio_lines) if ratio_lines else "계산 불가 (데이터 부족)"
    return fin_table, ratio_table


def generate_manual_deal_memo(req: ManualFinReq) -> str | None:
    if not check_claude_cli(): return None
    fin_table, ratio_table = _fmt_manual_fin(req)
    prompt = MANUAL_DEAL_MEMO_PROMPT.format(
        corp_name=req.corp_name, industry=req.industry or "미분류",
        deal_type=req.deal_type, year=req.year or "N/A",
        fin_table=fin_table, ratio_table=ratio_table, notes=req.notes or "없음",
    )
    return call_claude_cli(prompt, timeout=120)


def _manual_deal_memo_mock(req: ManualFinReq) -> str:
    fin_table, ratio_table = _fmt_manual_fin(req)
    has_fin = any(v is not None for v in [req.revenue, req.operating_profit, req.net_income])
    return f"""# Deal Memo — {req.corp_name} ({req.deal_type})

> ⚠️ Claude CLI 미연결 상태 — 기본 템플릿으로 생성됨. Claude CLI 연결 시 AI 기반 분석 자동 생성.

## 1. 기업 개요
- **기업명**: {req.corp_name}
- **산업/섹터**: {req.industry or '미분류'}
- **기준 연도**: {req.year or 'N/A'}
- **데이터 출처**: 수동 입력 (비상장/비공시 기업)

## 2. 재무 하이라이트
{fin_table if has_fin else '재무 데이터 미입력'}

## 3. 재무 건전성 평가
{ratio_table}

## 4. 투자/딜 관점 ({req.deal_type})
- 비상장 기업으로 DART 공시 데이터 부재
- 수동 입력 재무 데이터 기반 초기 검토 수준
- 추가 실사(DD) 통해 데이터 검증 필요

## 5. 핵심 리스크
- 공시 데이터 부재로 인한 정보 비대칭
- 감사 재무제표 미확보 가능성
- 오너 리스크 및 지배구조 불투명성
- 시장성 있는 비교군 확보 어려움

## 6. DD 체크리스트
- [ ] 최근 3개년 감사 재무제표 확보
- [ ] 주주 구성 및 지배구조 확인
- [ ] 핵심 고객사 및 매출 집중도 분석
- [ ] 주요 계약 및 IP 현황 검토
- [ ] 법적 분쟁 및 우발채무 확인
- [ ] 경영진 이력 및 Key Man 리스크 평가

## 7. 결론 및 권고
**판단**: Caution ⚠️
**근거**: 비상장 기업으로 공시 데이터 부재. 추가 DD를 통한 데이터 검증 후 본격 검토 권장.

---
*⚠️ 본 메모는 수동 입력 데이터 기반으로 작성되었으며, 투자 권유가 아닙니다.*
"""


@app.post("/api/manual-deal-memo")
async def manual_deal_memo(body: ManualFinReq):
    loop = asyncio.get_event_loop()
    def _run():
        ratios    = _calc_ratios_manual(body)
        memo_text = generate_manual_deal_memo(body)
        if not memo_text:
            memo_text = _manual_deal_memo_mock(body)
        memo_dir  = os.path.join(tempfile.gettempdir(), "memos")
        os.makedirs(memo_dir, exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_nm  = re.sub(r'[^\w가-힣-]', '_', body.corp_name)
        filename = f"{safe_nm}_{body.deal_type}_{ts}.md"
        with open(os.path.join(memo_dir, filename), "w", encoding="utf-8") as fh:
            fh.write(memo_text)
        return {"memo": memo_text, "filename": filename, "ratios": ratios}
    return await loop.run_in_executor(_pool, _run)

# ── 산업 인텔리전스: Pydantic Models ──────────────────────────────────────────
class CompanyTrendsReq(BaseModel):
    corp_name:  str
    industry:   str  = ""
    text_input: str  = ""
    yearly:     dict = {}
    scope:      str  = "domestic"   # "domestic" | "global" | "global_first"

class IndustryAnalysisReq(BaseModel):
    industry:   str
    text_input: str = ""
    scope:      str = "domestic"    # "domestic" | "global" | "global_first"


# ── 산업 인텔리전스: Endpoints ─────────────────────────────────────────────────
# ── 영문 기사 한글 번역 ──────────────────────────────────────────────────────────
TRANSLATE_PROMPT = """당신은 산업/딜/금융 뉴스 번역 에디터입니다.

## 역할
영문 뉴스 기사의 제목(title)과 요약(summary)을 한국어로 번역합니다.
독자는 산업 분석가 및 딜 인텔리전스 실무자이므로, 정확성과 가독성을 모두 갖춰야 합니다.

## 번역 지침
1. 최종 결과는 반드시 한국어로 작성합니다
2. 제목은 간결하고 자연스럽게 번역합니다 (직역보다 의미 전달 우선, 단 과도한 의역 금지)
3. 요약은 원문의 핵심 내용을 2~4문장 내로 자연스럽게 옮깁니다
4. 기업명·브랜드명·제품명·고유 기술 용어는 한글 표기 후 괄호 안에 원문 병기합니다
   예: 엔비디아(NVIDIA), 고대역폭메모리(HBM), 생성형 AI(Generative AI)
5. 원문에 없는 사실이나 수치를 추가하지 않습니다
6. 과장·선정적 표현(예: "폭등", "충격", "대박")은 사용하지 않습니다
7. 투자 추천·매수/매도 의견처럼 해석될 수 있는 표현은 금지합니다
8. 출처(source) 및 섹터(sector) 정보가 있으면 번역 톤 참고에만 활용하고, 출력에 포함하지 않습니다

## 입력 형식
아래 JSON 배열의 각 항목을 번역합니다:
{articles_json}

## 출력 형식
반드시 아래 JSON 배열 형식만 출력합니다. 설명·주석·마크다운 없이 순수 JSON만:
[
  {{
    "idx": <원본 idx 값>,
    "translated_title": "<한국어 제목>",
    "translated_summary": "<한국어 요약>"
  }},
  ...
]
"""


def translate_articles_if_needed(articles: list) -> list:
    """source_language == 'en' 기사의 title/description을 한글로 번역."""
    en_idx = [i for i, a in enumerate(articles) if a.get("source_language") == "en"
              and not a.get("translated_title")]
    if not en_idx or not check_claude_cli():
        for i in en_idx:
            articles[i]["translation_status"] = "skipped"
        return articles

    batch = [
        {
            "idx": i,
            "title": articles[i].get("title", ""),
            "summary": articles[i].get("description", ""),
            "source": articles[i].get("source", ""),
            "sector": articles[i].get("sector", ""),
        }
        for i in en_idx[:5]
    ]
    try:
        import json as _json
        prompt = TRANSLATE_PROMPT.format(articles_json=_json.dumps(batch, ensure_ascii=False))
        response = call_claude_cli(prompt, timeout=150)
        if not response:
            raise ValueError("empty response")
        parsed = _extract_json(response)
        if parsed is None:
            raise ValueError(f"JSON 파싱 실패: {response[:200]}")
        if not isinstance(parsed, list):
            parsed = parsed.get("translations") or parsed.get("items") or [parsed]
        for item in (parsed or []):
            idx = item.get("idx")
            if idx is not None and idx < len(articles):
                articles[idx]["translated_title"]   = item.get("translated_title", "")
                articles[idx]["translated_summary"]  = item.get("translated_summary", "")
                articles[idx]["translation_status"]  = "ok"
    except Exception:
        for i in en_idx:
            articles[i]["translation_status"] = "failed"
    return articles


def _normalize_company_result(raw: dict, news: list) -> dict:
    """Claude가 구 형식으로 응답할 때 신 형식으로 변환."""
    def _to_issues(items):
        if not items: return []
        if isinstance(items[0], dict): return items
        return [{"title": str(x)[:30], "summary": ""} for x in items]

    if not raw.get("domestic_issues"):
        src = raw.get("recent_issues") or []
        if not src and raw.get("core_signals"):
            src = [{"title": s[:30], "summary": ""} for s in raw["core_signals"]]
        raw["domestic_issues"] = _to_issues(src)

    if not raw.get("global_issues"):
        en_news = [n for n in news if n.get("source_language") == "en"]
        if en_news:
            raw["global_issues"] = [
                {
                    "title":          n.get("translated_title") or n["title"],
                    "summary":        n.get("translated_summary") or (n.get("description") or ""),
                    "original_title": n["title"] if n.get("translated_title") else None,
                    "is_translated":  bool(n.get("translated_title")),
                }
                for n in en_news[:4]
            ]
        else:
            raw["global_issues"] = []

    if not raw.get("common_signals"):
        raw["common_signals"] = (
            raw.get("core_signals") or
            raw.get("recent_trends") or []
        )[:3]

    if not raw.get("deal_angle"):
        ops = raw.get("deal_opportunities") or []
        raw["deal_angle"] = ops[0] if ops else ""

    return raw


def _normalize_industry_result(raw: dict, news: list) -> dict:
    """Claude가 구 형식으로 응답할 때 신 형식으로 변환."""
    def _to_issues(items):
        if not items: return []
        if isinstance(items[0], dict): return items
        return [{"title": str(x)[:30], "summary": ""} for x in items]

    if not raw.get("domestic_issues"):
        src = raw.get("recent_trends") or raw.get("risks") or []
        raw["domestic_issues"] = _to_issues(src)

    if not raw.get("global_issues"):
        en_news = [n for n in news if n.get("source_language") == "en"]
        if en_news:
            raw["global_issues"] = [
                {
                    "title":          n.get("translated_title") or n["title"],
                    "summary":        n.get("translated_summary") or (n.get("description") or ""),
                    "original_title": n["title"] if n.get("translated_title") else None,
                    "is_translated":  bool(n.get("translated_title")),
                }
                for n in en_news[:4]
            ]
        else:
            raw["global_issues"] = []

    if not raw.get("common_signals"):
        raw["common_signals"] = (
            raw.get("growth_drivers") or
            raw.get("recent_trends") or []
        )[:3]

    if not raw.get("deal_angle"):
        ops = raw.get("deal_opportunities") or []
        raw["deal_angle"] = ops[0] if ops else ""

    return raw



# ── Deal Intelligence ──────────────────────────────────────────────────────────
_DEAL_KW_KO = [
    '인수', '합병', '매각', '지분', '투자유치', '투자 유치', '시리즈', '프리IPO',
    '상장', 'IPO', '기업공개', '사모펀드', 'PEF', '경영권', '유상증자', 'CB',
    'BW', '메자닌', '리파이낸싱', '구조조정', '회생', '워크아웃', 'JV', '합작법인',
    '합작회사', '전략적 제휴', '기술이전', '라이선스아웃', '라이선스 아웃',
    '공급계약', '장기계약', '컨소시엄', '엑시트', '바이아웃', '지분취득',
    '분할매각', '인수합병', 'MOU', '업무협약', '수주', '대규모 계약',
]
_DEAL_KW_EN = [
    'M&A', 'merger', 'acquisition', 'stake sale', 'buyout', 'private equity',
    'venture capital', 'Series A', 'Series B', 'Series C', 'Series D',
    'IPO', 'pre-IPO', 'going public', 'strategic investment', 'fundraising',
    'funding round', 'convertible bond', 'mezzanine', 'restructuring',
    'divestiture', 'spin-off', 'spinoff', 'joint venture', 'licensing deal',
    'license out', 'strategic alliance', 'carve-out', 'LBO', 'MBO',
    'rights offering', 'equity raise', 'takeover', 'investment round',
    'tender offer', 'anchor investor', 'deal', 'partnership',
]
_DEAL_TYPE_KO = [
    ('M&A',                  ['인수', '합병', '경영권', '인수합병', '피인수', '지분 취득', '지분취득']),
    ('VC 투자유치',           ['시리즈', '투자유치', '투자 유치', '벤처투자']),
    ('PE / 바이아웃',         ['사모펀드', 'PEF', '바이아웃', '경영참여']),
    ('IPO / Pre-IPO',        ['상장', 'IPO', '기업공개', '프리IPO']),
    ('자본조달',              ['유상증자', 'CB', 'BW', '메자닌', '리파이낸싱', '전환사채']),
    ('구조조정',              ['구조조정', '회생', '워크아웃', '분할매각']),
    ('JV / 전략적 제휴',     ['JV', '합작', '전략적 제휴', 'MOU', '업무협약']),
    ('기술이전 / 라이선스',  ['기술이전', '라이선스아웃', '라이선스 아웃']),
    ('대규모 공급계약',       ['공급계약', '장기계약', '수주', '대규모 계약']),
    ('매각',                  ['매각']),
]
_DEAL_TYPE_EN = [
    ('M&A',                  ['merger', 'acquisition', 'takeover', 'buyout', 'stake']),
    ('VC 투자유치',           ['Series A', 'Series B', 'Series C', 'Series D', 'funding round', 'venture capital']),
    ('PE / 바이아웃',         ['private equity', 'LBO', 'MBO', 'leveraged buyout']),
    ('IPO / Pre-IPO',        ['IPO', 'pre-IPO', 'going public', 'listing']),
    ('자본조달',              ['convertible bond', 'mezzanine', 'rights offering', 'equity raise']),
    ('구조조정',              ['restructuring', 'divestiture', 'spin-off', 'spinoff', 'carve-out']),
    ('JV / 전략적 제휴',     ['joint venture', 'strategic alliance', 'partnership']),
    ('기술이전 / 라이선스',  ['licensing deal', 'license out']),
    ('대규모 공급계약',       ['supply agreement', 'long-term contract']),
]

# 점수 기준 임계값: score >= DEAL_SCORE_THRESHOLD 인 경우만 deal_related=True
DEAL_SCORE_THRESHOLD = 3

def classify_deal_relevance(article: dict) -> dict:
    """키워드 규칙 기반 딜 관련성 분류. is_deal_related, deal_type, deal_relevance_score 추가."""
    text = ' '.join(filter(None, [
        article.get('title', ''),
        article.get('description', ''),
        article.get('translated_title', ''),
        article.get('translated_summary', ''),
    ])).lower()
    lang = article.get('source_language', 'ko')
    kw_list  = _DEAL_KW_KO if lang == 'ko' else _DEAL_KW_EN
    rules    = _DEAL_TYPE_KO if lang == 'ko' else _DEAL_TYPE_EN

    matched = [kw for kw in kw_list if kw.lower() in text]
    count   = len(matched)
    score   = 5 if count >= 4 else 4 if count == 3 else 3 if count >= 1 else 1

    deal_type = '기타 Deal Issue'
    for dtype, kws in rules:
        if any(k.lower() in text for k in kws):
            deal_type = dtype
            break

    is_deal = score >= DEAL_SCORE_THRESHOLD
    return {
        **article,
        'is_deal_related':       is_deal,
        'deal_type':             deal_type if is_deal else '',
        'deal_relevance_score':  score,
        'deal_matched_keywords': matched[:5],
        'why_deal_related': (
            f"키워드 매칭 (검색 결과 요약 기준): {', '.join(matched[:3])}" if matched else "일반 산업 뉴스"
        ),
        'deal_implication':   '',
        'related_companies':  [],
        'next_checkpoints':   [],
    }

def extract_deal_highlights(articles: list) -> list:
    """기사 목록에서 딜 관련 기사만 추출·정렬."""
    classified = [classify_deal_relevance(a) for a in articles]
    deals = [a for a in classified if a.get('is_deal_related')]
    deals.sort(key=lambda x: x.get('deal_relevance_score', 0), reverse=True)
    return deals

DEAL_HIGHLIGHT_PROMPT = """당신은 딜 인텔리전스 애널리스트입니다.

아래 뉴스 기사들은 {sector} 산업 관련 딜 이슈로 분류된 기사입니다.
각 기사에 대해 딜 관점 분석을 제공하세요.

주의:
- 기사 제목과 요약만 기반으로 판단합니다
- 확인되지 않은 딜 규모·거래 상대방·조건은 임의로 생성하지 마세요
- 추정·가능성은 반드시 "추가 확인 필요" 또는 "~로 추정됨"으로 표현하세요
- 원문을 읽지 않은 기사라면 "검색 결과 요약 기준" 표기를 포함하세요

기사 목록 (JSON):
{articles_json}

출력: 아래 JSON 배열 형식만, 설명 없이:
[
  {{
    "idx": 0,
    "deal_type": "M&A / VC 투자유치 / PE 바이아웃 / IPO 및 Pre-IPO / 자본조달 / 구조조정 / JV 및 전략적 제휴 / 기술이전 및 라이선스 / 대규모 공급계약 / 기타",
    "deal_relevance_score": 1~5,
    "why_deal_related": "한 줄 이유 (검색 결과 요약 기준 명시 포함)",
    "deal_implication": "딜 관점 시사점 2~3문장 (추정은 ~로 추정됨 표기)",
    "related_companies": ["회사명1"],
    "next_checkpoints": ["추가 확인사항1", "추가 확인사항2"]
  }}
]
"""

def generate_deal_highlights_claude(deal_articles: list, sector: str = '') -> list:
    """Claude CLI로 딜 하이라이트 분석. 실패 시 키워드 기반 결과 반환."""
    if not deal_articles or not check_claude_cli():
        return deal_articles
    batch = [
        {
            'idx': i,
            'title': a.get('translated_title') or a.get('title', ''),
            'description': a.get('translated_summary') or a.get('description', ''),
            'source': a.get('source', ''),
            'url': a.get('url', ''),
        }
        for i, a in enumerate(deal_articles[:10])  # 최대 10건
    ]
    try:
        prompt = DEAL_HIGHLIGHT_PROMPT.format(
            sector=sector or '해당 산업',
            articles_json=json.dumps(batch, ensure_ascii=False)
        )
        response = call_claude_cli(prompt, timeout=90)
        if not response:
            return deal_articles
        parsed = _extract_json(response)
        if not isinstance(parsed, list):
            return deal_articles
        for item in parsed:
            idx = item.get('idx')
            if idx is not None and idx < len(deal_articles):
                deal_articles[idx]['deal_type']            = item.get('deal_type', deal_articles[idx]['deal_type'])
                deal_articles[idx]['deal_relevance_score'] = item.get('deal_relevance_score', deal_articles[idx]['deal_relevance_score'])
                deal_articles[idx]['why_deal_related']     = item.get('why_deal_related', deal_articles[idx]['why_deal_related'])
                deal_articles[idx]['deal_implication']     = item.get('deal_implication', '')
                deal_articles[idx]['related_companies']    = item.get('related_companies', [])
                deal_articles[idx]['next_checkpoints']     = item.get('next_checkpoints', [])
    except Exception:
        pass  # 키워드 기반 결과 그대로 사용
    return deal_articles


class MacroDealReq(BaseModel):
    korea_macro: str = ""          # 한국 매크로 지표 (자유 텍스트)
    global_macro: str = ""         # 글로벌 매크로 지표
    market_indicators: str = ""    # 시장지표
    sector: str = ""               # 선택 산업
    company: str = ""              # 선택 기업명
    dart_result: str = ""          # DART 재무분석 결과 (선택)
    industry_intel: str = ""       # 산업 인텔리전스 결과 (선택)
    report_date: str = ""          # 분석 기준일 (미입력 시 오늘)


class ArticleDealMemoReq(BaseModel):
    title:         str = ""
    summary:       str = ""
    url:           str = ""
    deal_type:     str = ""
    deal_implication: str = ""
    company_name:  str = ""
    sector:        str = ""

ARTICLE_DEAL_MEMO_PROMPT = """당신은 딜 메모 작성 전문가입니다.

아래 딜 이슈 기사를 기반으로 간략한 Deal Note를 작성하세요.

기사 정보:
- 제목: {title}
- 요약: {summary}
- 딜 유형: {deal_type}
- 딜 시사점: {deal_implication}
- 관련 기업: {company_name}
- 관련 산업: {sector}
- 출처 URL: {url}

작성 지침:
- 확인되지 않은 사실(딜 규모, 가격, 조건 등)은 절대 임의로 작성하지 마세요
- 추정 내용은 반드시 "추가 확인 필요" 또는 "~로 알려짐" 등으로 표기
- 기사 제목/요약 기준으로 작성하는 것임을 명시
- 분량: 300~500자 내외 한국어 Markdown

출력 형식 (Markdown):
## Deal Note — {deal_type}

### 배경
[딜 배경 및 개요]

### 주요 내용
[알려진 사실 위주, 추정 명시]

### 딜 관점 시사점
[투자·거래 관점 분석]

### 추가 확인사항
- [ ] 항목1
- [ ] 항목2

> 출처: {url}
> 작성 기준: 검색 결과 요약 기준 (기사 원문 미열람)
"""

MACRO_DEAL_PROMPT = """당신은 VC, PE, M&A, IPO, CFO 관점에서 거시경제와 시장지표를 해석하는 Macro Deal Intelligence Analyst입니다.

## 원칙
- 입력 수치가 있는 경우에만 구체적 판단을 내립니다. 수치가 없으면 "추가 확인 필요"로 표시합니다.
- 데이터 기준일을 반드시 명시합니다. 불확실한 인과관계는 "~일 가능성", "~으로 판단됨" 등 유보적 표현을 씁니다.
- 투자 추천·매수/매도 의견으로 해석될 수 있는 표현은 사용하지 않습니다.
- 최종 결과는 한국어로 작성합니다.

## 입력 데이터

[한국 매크로]
{korea_macro}

[글로벌 매크로]
{global_macro}

[시장지표 — 주가지수 / 금리 / 환율 / 원자재 / 스프레드]
{market_indicators}

[분석 대상 산업]
{sector}

[분석 대상 기업]
{company}

[DART 재무분석 결과 (선택)]
{dart_result}

[산업 인텔리전스 결과 (선택)]
{industry_intel}

---

## 출력 지시

분석 기준일 **{report_date}** 기준으로, 아래 10개 섹션을 마크다운으로 작성합니다.
각 섹션은 핵심 인사이트를 3~6줄 이내 bullet로 서술합니다. 해당 데이터가 없는 항목은 "입력 데이터 없음"으로 표시합니다.

# Macro Deal Intelligence Report
> 분석 기준일: {report_date} | 산업: {sector_label} | 기업: {company_label}

## 1. Executive Summary
현재 매크로 국면(Risk-On / Risk-Off / 혼조)과 딜 실행 관점 핵심 메시지 3~5줄.

## 2. 금리 환경
기준금리·장단기 금리차·회사채 스프레드가 밸류에이션 할인율, 차입 비용, LBO 여건, 리파이낸싱 리스크에 미치는 영향.

## 3. FX 환경
원/달러 및 주요 통화 환율이 매출 환산, 수입 원자재 원가, 외화부채 평가손익, Cross-border M&A 가격에 미치는 영향.

## 4. 물가 및 원가 압력
CPI·PPI·원자재 가격이 COGS율, EBITDA 마진, 가격 전가력에 미치는 영향.

## 5. 경기 및 수요 전망
GDP·PMI·소비자신뢰지수가 탑라인 성장성, 재고·매출채권 리스크, 투자 thesis 유효성에 미치는 영향.

## 6. 유동성 및 딜 마켓
M2·신용 스프레드·IPO 온도가 VC 투자, M&A 멀티플, IPO 창구, 사모대출, Secondary 환경에 미치는 영향.

## 7. 산업 민감도 ({sector_label})
해당 산업이 어떤 매크로 변수(금리·환율·원자재·경기)에 가장 민감한지, 현재 국면이 유리한지 불리한지 서술. 민감도 표(변수 | 민감도 | 현재 국면 | 평가) 포함.

## 8. 밸류에이션 함의 ({company_label})
현 매크로 기준 EV/EBITDA·PER·EV/Sales 멀티플 방향성. DART 데이터 연계 시 해당 기업 재무 영향 서술.

## 9. 재무실사 체크리스트
매크로 변수 변화에 따라 DD 시 반드시 확인할 항목을 체크리스트([ ]) 형식으로 작성.
(금리·FX·원가·경기·유동성·공시 축별 최소 1개 이상)

## 10. CFO 액션 아이템
즉시(0~1개월) / 단기(1~3개월) / 중기(3~12개월) 구분하여 CFO가 점검해야 할 액션 아이템 서술.

---
*본 분석은 입력 데이터 기준이며 투자 추천을 포함하지 않습니다.*
"""


@app.post("/api/article-deal-memo")
async def article_deal_memo(body: ArticleDealMemoReq):
    loop = asyncio.get_event_loop()
    def _run():
        if not check_claude_cli():
            return {"memo": f"## Deal Note — {body.deal_type or '딜 이슈'}\n\n> Claude CLI 미연결로 메모를 생성할 수 없습니다.\n\n**기사 제목**: {body.title}\n\n**딜 시사점**: {body.deal_implication or '(분석 없음)'}\n\n> 출처: {body.url}", "ok": False}
        prompt = ARTICLE_DEAL_MEMO_PROMPT.format(
            title=body.title, summary=body.summary, url=body.url,
            deal_type=body.deal_type, deal_implication=body.deal_implication,
            company_name=body.company_name, sector=body.sector,
        )
        memo = call_claude_cli(prompt, timeout=120)
        return {"memo": memo or "(생성 실패)", "ok": bool(memo)}
    return await loop.run_in_executor(_pool, _run)



# ══════════════════════════════════════════════════════════════
# MACRO MODULE — Catalog · Mock · Cache · Fetch · Endpoints
# ══════════════════════════════════════════════════════════════

MACRO_CATALOG = {
    "domestic": [
        {"id":"kr_rate",    "name":"기준금리",        "unit":"%",       "category":"금리"},
        {"id":"kr_3y",      "name":"국고채 3년",      "unit":"%",       "category":"금리"},
        {"id":"kr_10y",     "name":"국고채 10년",     "unit":"%",       "category":"금리"},
        {"id":"usdkrw",     "name":"원/달러 환율",    "unit":"원",      "category":"환율"},
        {"id":"kr_cpi",     "name":"CPI",             "unit":"%(YoY)", "category":"물가"},
        {"id":"kr_ppi",     "name":"PPI",             "unit":"%(YoY)", "category":"물가"},
        {"id":"kr_gdp",     "name":"GDP 성장률",      "unit":"%(YoY)", "category":"경기"},
        {"id":"kr_export",  "name":"수출",            "unit":"억달러",  "category":"무역"},
        {"id":"kr_import",  "name":"수입",            "unit":"억달러",  "category":"무역"},
        {"id":"kr_ca",      "name":"경상수지",        "unit":"억달러",  "category":"무역"},
        {"id":"kr_csi",     "name":"소비자심리지수",  "unit":"pt",      "category":"심리"},
        {"id":"kr_bsi",     "name":"기업경기 BSI",    "unit":"pt",      "category":"심리"},
        {"id":"kr_unemploy","name":"실업률",          "unit":"%",       "category":"고용"},
        {"id":"kr_hpi",     "name":"주택가격지수",    "unit":"%(MoM)", "category":"부동산"},
    ],
    "global": [
        {"id":"us_rate",    "name":"미국 기준금리",   "unit":"%",        "category":"금리"},
        {"id":"us_10y",     "name":"미국 10년물",     "unit":"%",        "category":"금리"},
        {"id":"us_cpi",     "name":"미국 CPI",        "unit":"%(YoY)",  "category":"물가"},
        {"id":"us_gdp",     "name":"미국 GDP",        "unit":"%(QoQ*)", "category":"경기"},
        {"id":"us_unemploy","name":"미국 실업률",     "unit":"%",        "category":"고용"},
        {"id":"dxy",        "name":"달러인덱스",      "unit":"pt",       "category":"환율"},
        {"id":"wti",        "name":"WTI 유가",        "unit":"USD/bbl",  "category":"원자재"},
        {"id":"brent",      "name":"브렌트유",        "unit":"USD/bbl",  "category":"원자재"},
        {"id":"copper",     "name":"구리 가격",       "unit":"USD/ton",  "category":"원자재"},
        {"id":"gold",       "name":"금 가격",         "unit":"USD/oz",   "category":"원자재"},
        {"id":"vix",        "name":"VIX",             "unit":"pt",       "category":"심리"},
    ],
    "market": [
        {"id":"kospi",  "name":"KOSPI",         "unit":"pt",       "category":"주가"},
        {"id":"kosdaq", "name":"KOSDAQ",        "unit":"pt",       "category":"주가"},
        {"id":"sp500",  "name":"S&P 500",       "unit":"pt",       "category":"주가"},
        {"id":"nasdaq", "name":"NASDAQ",        "unit":"pt",       "category":"주가"},
        {"id":"vix",    "name":"VIX",           "unit":"pt",       "category":"심리"},
        {"id":"dxy",    "name":"달러인덱스",    "unit":"pt",       "category":"환율"},
        {"id":"usdkrw", "name":"원/달러 환율",  "unit":"원",       "category":"환율"},
        {"id":"us_10y", "name":"미국 10년물",   "unit":"%",        "category":"금리"},
        {"id":"wti",    "name":"WTI 유가",      "unit":"USD/bbl",  "category":"원자재"},
    ],
}

_DEFAULT_IDS = {
    "domestic": ["kr_rate","kr_3y","usdkrw","kr_cpi","kr_gdp","kr_export","kr_csi"],
    "global":   ["us_rate","us_10y","us_cpi","us_gdp","dxy","wti","vix"],
    "market":   ["kospi","sp500","usdkrw","vix"],
}

# mock data — realistic May 2025 reference values
_MACRO_MOCK: dict = {
    "kr_rate":    {"value":3.50,  "date":"2025-04","mom": 0.00,"yoy":-0.25,"trend":"→","note":"2025.4 동결"},
    "kr_3y":      {"value":2.85,  "date":"2025-04","mom":-0.05,"yoy":-0.40,"trend":"↓"},
    "kr_10y":     {"value":3.05,  "date":"2025-04","mom":-0.03,"yoy":-0.35,"trend":"↓"},
    "usdkrw":     {"value":1375,  "date":"2025-04","mom":-5,   "yoy": 15,  "trend":"→"},
    "kr_cpi":     {"value":2.1,   "date":"2025-03","mom":-0.1, "yoy":-0.5, "trend":"↓"},
    "kr_ppi":     {"value":1.8,   "date":"2025-03","mom": 0.2, "yoy":-1.0, "trend":"→"},
    "kr_gdp":     {"value":1.5,   "date":"2025-Q1","mom": None,"yoy":-0.7, "trend":"↓"},
    "kr_export":  {"value":572,   "date":"2025-04","mom": 12,  "yoy": 32,  "trend":"↑"},
    "kr_import":  {"value":541,   "date":"2025-04","mom":  8,  "yoy": 18,  "trend":"↑"},
    "kr_ca":      {"value":38,    "date":"2025-03","mom":  5,  "yoy": 12,  "trend":"↑"},
    "kr_csi":     {"value":97.1,  "date":"2025-04","mom":-1.3, "yoy":-3.2, "trend":"↓"},
    "kr_bsi":     {"value":74,    "date":"2025-04","mom":-2,   "yoy":-6,   "trend":"↓"},
    "kr_unemploy":{"value":2.8,   "date":"2025-03","mom": 0.1, "yoy": 0.2, "trend":"→"},
    "kr_hpi":     {"value":-0.1,  "date":"2025-03","mom":-0.1, "yoy":-1.5, "trend":"↓"},
    "us_rate":    {"value":4.375, "date":"2025-04","mom": 0.0, "yoy":-0.75,"trend":"→","note":"4.25~4.50%"},
    "us_10y":     {"value":4.35,  "date":"2025-04","mom":-0.10,"yoy":-0.05,"trend":"→"},
    "us_cpi":     {"value":2.4,   "date":"2025-03","mom":-0.1, "yoy":-0.6, "trend":"↓"},
    "us_gdp":     {"value":2.7,   "date":"2025-Q4","mom": None,"yoy": 0.2, "trend":"→"},
    "us_unemploy":{"value":4.2,   "date":"2025-03","mom": 0.1, "yoy": 0.4, "trend":"→"},
    "dxy":        {"value":100.5, "date":"2025-04","mom":-1.5, "yoy":-3.2, "trend":"↓"},
    "wti":        {"value":70.2,  "date":"2025-04","mom":-3.1, "yoy":-8.5, "trend":"↓"},
    "brent":      {"value":73.8,  "date":"2025-04","mom":-3.0, "yoy":-7.8, "trend":"↓"},
    "copper":     {"value":9950,  "date":"2025-04","mom": 150, "yoy": 820, "trend":"↑"},
    "gold":       {"value":3300,  "date":"2025-04","mom": 150, "yoy": 680, "trend":"↑"},
    "vix":        {"value":22.5,  "date":"2025-04","mom": 3.2, "yoy": 5.1, "trend":"↑"},
    "kospi":      {"value":2560,  "date":"2025-04","mom": 45,  "yoy": 82,  "trend":"↑"},
    "kosdaq":     {"value":750,   "date":"2025-04","mom": 12,  "yoy":-15,  "trend":"→"},
    "sp500":      {"value":5780,  "date":"2025-04","mom": 120, "yoy": 650, "trend":"↑"},
    "nasdaq":     {"value":18350, "date":"2025-04","mom": 380, "yoy":2200, "trend":"↑"},
}

# ECOS series map: id → (stat_code, item_code, period)
_ECOS_MAP: dict = {
    "kr_rate":    ("722Y001","0101000","M"),
    "kr_3y":      ("817Y002","010190000","M"),
    "kr_10y":     ("817Y002","010200000","M"),
    "usdkrw":     ("731Y001","0000001","D"),
    "kr_cpi":     ("901Y009","0","M"),
    "kr_ppi":     ("404Y014","P0","M"),
    "kr_gdp":     ("200Y001","10000","Q"),
    "kr_export":  ("301Y013","Z","M"),
    "kr_import":  ("301Y013","D","M"),
    "kr_ca":      ("301Y016","00","M"),
    "kr_csi":     ("512Y003","99988","M"),
    "kr_bsi":     ("512Y007","BSI0001","M"),
    "kr_unemploy":("901Y027","000","M"),
}

# Indicators that need units=pc1 (percent change from year ago) from FRED.
# CPIAUCSL returns index level (~330) — units=pc1 returns actual YoY% (~2-3%).
_FRED_PC1_INDICATORS: set = {"us_cpi"}

# ECOS indicators that return index levels but are labeled %(YoY).
# We fetch 14 months and compute (current/12ago - 1)*100 ourselves.
_ECOS_YOY_INDICATORS: set = {"kr_cpi", "kr_ppi"}

# FRED series map: id → series_id
_FRED_MAP: dict = {
    "us_rate":    "FEDFUNDS",
    "us_10y":     "DGS10",
    "us_cpi":     "CPIAUCSL",
    "us_gdp":     "A191RL1Q225SBEA",
    "us_unemploy":"UNRATE",
    "dxy":        "DTWEXBGS",
    "wti":        "DCOILWTICO",
    "brent":      "DCOILBRENTEU",
    "gold":       "GOLDPMGBD228NLBM",
    "vix":        "VIXCLS",
    "sp500":      "SP500",
    "nasdaq":     "NASDAQCOM",
    "copper":     "PCOPPUSDM",
}

# 무료 실시간 소스 없는 지표 (mock 데이터로만 제공)
_NO_REALTIME: set = {"kospi", "kosdaq", "gold"}


def _macro_cache_get(indicator_id: str) -> dict | None:
    """24h 이내 캐시 반환, 없으면 None"""
    with _db() as conn:
        row = conn.execute(
            "SELECT value_json, fetched_at FROM macro_cache WHERE indicator_id=? "
            "AND datetime(fetched_at) > datetime('now','-24 hours')",
            (indicator_id,)
        ).fetchone()
    if row:
        import json as _j
        return _j.loads(row[0])
    return None


def _macro_cache_set(indicator_id: str, data: dict) -> None:
    import json as _j
    with _db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO macro_cache(indicator_id,value_json,fetched_at) "
            "VALUES(?,?,datetime('now'))",
            (indicator_id, _j.dumps(data, ensure_ascii=False))
        )


def _fetch_ecos(indicator_id: str, api_key: str) -> dict | None:
    from datetime import date as _d
    import requests as _r, json as _j
    if indicator_id not in _ECOS_MAP:
        return None
    stat, item, period = _ECOS_MAP[indicator_id]
    today = _d.today()
    need_yoy = indicator_id in _ECOS_YOY_INDICATORS
    limit = 14 if need_yoy else 5
    if period == "Q":
        start = f"{today.year-1}Q1"; end = f"{today.year}Q2"
    elif period == "D":
        from datetime import timedelta as _td
        d_start = today - _td(days=30)
        start = d_start.strftime("%Y%m%d"); end = today.strftime("%Y%m%d")
    else:
        # For YoY indicators fetch 2 years back to ensure 13+ months
        start = (f"{(today.year-2)*100+1:06d}" if need_yoy
                 else f"{(today.year-1)*100+today.month:06d}")
        end = f"{today.year*100+today.month:06d}"
    url = (f"https://ecos.bok.or.kr/api/StatisticSearch/{api_key}/json/kr"
           f"/1/{limit}/{stat}/{period}/{start}/{end}/{item}")
    try:
        resp = _r.get(url, timeout=8)
        rows = resp.json().get("StatisticSearch",{}).get("row",[])
        if not rows:
            return None
        rows.sort(key=lambda x: x.get("TIME",""), reverse=True)
        latest = rows[0]; prev = rows[1] if len(rows)>1 else None
        v = float(latest.get("DATA_VALUE","0").replace(",",""))
        pv = float(prev.get("DATA_VALUE","0").replace(",","")) if prev else None
        if need_yoy and len(rows) >= 13:
            # rows[0]=current month, rows[12]=same month last year
            v_12ago = float(rows[12].get("DATA_VALUE","0").replace(",",""))
            value = round((v / v_12ago - 1) * 100, 2) if v_12ago else None
            mom_pct = round((v / pv - 1) * 100, 4) if pv else None
            trend = "↑" if (value or 0) > 0 else ("↓" if (value or 0) < 0 else "→")
            return {"value": value, "date": latest.get("TIME",""),
                    "mom": mom_pct, "yoy": None, "trend": trend, "source": "ECOS"}
        mom = round(v - pv, 4) if pv is not None else None
        return {"value": v, "date": latest.get("TIME",""), "mom": mom,
                "yoy": None, "trend": "→", "source": "ECOS"}
    except Exception:
        return None


def _fetch_fred(indicator_id: str, api_key: str) -> dict | None:
    import requests as _r
    if indicator_id not in _FRED_MAP:
        return None
    series = _FRED_MAP[indicator_id]
    use_pc1 = indicator_id in _FRED_PC1_INDICATORS
    units_param = "&units=pc1" if use_pc1 else ""
    url = (f"https://api.stlouisfed.org/fred/series/observations"
           f"?series_id={series}&api_key={api_key}&sort_order=desc"
           f"&limit=14{units_param}&file_type=json")
    try:
        resp = _r.get(url, timeout=8)
        obs = [o for o in resp.json().get("observations",[]) if o.get("value",".") != "."]
        if not obs:
            return None
        latest = obs[0]; prev = obs[1] if len(obs)>1 else None
        v = float(latest["value"])
        pv = float(prev["value"]) if prev else None
        mom = round(v - pv, 4) if pv is not None else None
        if use_pc1:
            # v is already the YoY% (FRED computed); no separate yoy needed
            yoy = None
        else:
            yoy_obs = obs[12] if len(obs) > 12 else None
            yv = float(yoy_obs["value"]) if yoy_obs else None
            yoy = round(v - yv, 4) if yv is not None else None
        trend = "↑" if (mom or 0) > 0 else ("↓" if (mom or 0) < 0 else "→")
        return {"value": v, "date": latest["date"], "mom": mom,
                "yoy": yoy, "trend": trend, "source": "FRED"}
    except Exception:
        return None


def _get_indicator(indicator_id: str, force: bool = False) -> dict:
    """캐시 → ECOS/FRED API → mock 순서로 데이터 반환"""
    # 실시간 소스가 없는 지표는 바로 mock 반환 (캐시 불필요)
    if indicator_id in _NO_REALTIME:
        mock = _MACRO_MOCK.get(indicator_id, {})
        return {**mock, "from_cache": False, "source": "no_source"}
    if not force:
        cached = _macro_cache_get(indicator_id)
        if cached:
            return {**cached, "from_cache": True}
    ecos_key = os.getenv("ECOS_API_KEY","")
    fred_key = os.getenv("FRED_API_KEY","")
    result = None
    if ecos_key and indicator_id in _ECOS_MAP:
        result = _fetch_ecos(indicator_id, ecos_key)
    if result is None and fred_key and indicator_id in _FRED_MAP:
        result = _fetch_fred(indicator_id, fred_key)
    if result:
        _macro_cache_set(indicator_id, result)
        return {**result, "from_cache": False}
    mock = _MACRO_MOCK.get(indicator_id, {})
    # has_mapping: API 키/네트워크 문제일 수 있음 / no mapping: 정말 소스 없음
    has_mapping = indicator_id in _ECOS_MAP or indicator_id in _FRED_MAP
    src = "mock" if has_mapping else "no_source"
    return {**mock, "from_cache": False, "source": src}


MACRO_SNAPSHOT_PROMPT = """당신은 VC, PE, M&A, IPO, CFO 관점으로 매크로 경제를 해석하는 Macro Deal Intelligence Analyst입니다.

## 원칙
- 수치가 있는 경우만 구체적으로 서술합니다. 없으면 "데이터 없음"으로 표시합니다.
- 각 섹션은 bullet 3~5개, 간결하게 작성합니다.
- 투자 추천·매수/매도 의견 표현은 사용하지 않습니다.
- 최종 결과는 한국어로 작성합니다.

## 입력

분석 기준일: {data_date}
분석 유형: {analysis_type}
분석 범위: {scope_label}
대상 산업: {sector}

지표 데이터:
{indicators_json}

---

## 출력 — 아래 섹션을 순서대로 마크다운으로 작성하세요

# Macro Snapshot
> {analysis_type} | {data_date} 기준

**현재 국면**: [Risk-On / Risk-Off / 혼조 중 선택, 근거 1줄]

- **금리 환경**: (금리 수준과 방향성, 딜 비용 영향 1줄)
- **환율 환경**: (원/달러 방향과 수출입·외화 영향 1줄)
- **물가 환경**: (인플레이션 수준과 원가 압력 1줄)
- **경기 환경**: (성장 사이클 국면 1줄)
- **유동성 환경**: (금리·VIX 기반 시장 유동성 1줄)
- **시장심리**: (VIX·주가 방향 기반 리스크 선호도 1줄)

---

# Domestic Macro Trend
입력된 국내 지표를 바탕으로 아래 항목 중 데이터가 있는 것만 서술합니다.

- **금리**: 
- **환율**: 
- **물가**: 
- **경기**: 
- **소비**: 
- **수출입**: 
- **금융시장**: 

---

# Global Macro Trend
입력된 글로벌 지표를 바탕으로 아래 항목 중 데이터가 있는 것만 서술합니다.

- **미국 금리**: 
- **미국 물가**: 
- **글로벌 경기**: 
- **원자재**: 
- **위험자산 선호**: 
- **달러 강세/약세**: 

---

# Deal Implication
현재 매크로 환경이 딜 생태계에 미치는 영향을 서술합니다.

- **기업 밸류에이션**: (멀티플 방향성)
- **자금조달**: (차입 비용, 인수금융 여건)
- **IPO**: (상장 창구 온도)
- **M&A**: (전략적/재무적 인수자 활동)
- **VC 투자**: (벤처 라운드 환경)
- **PE 투자**: (바이아웃 여건)
- **리파이낸싱**: (기존 부채 재조달 비용)
- **CFO 의사결정**: (단기 우선 과제)

---

# Sector Sensitivity
{sector_section}
"""


class MacroAnalyzeReq(BaseModel):
    scope: str = "both"                         # domestic / global / both
    selected_indicators: list[str] = []         # 빈 배열이면 기본 대표 지표 사용
    sector: str = ""
    company_name: str = ""
    use_default_if_empty: bool = True
    force_refresh: bool = False                 # True이면 캐시 무시하고 API 재조회



@app.post("/api/macro-deal-intelligence")
async def macro_deal_intelligence(body: MacroDealReq):
    loop = asyncio.get_event_loop()
    def _run():
        from datetime import date as _date
        report_date = body.report_date or _date.today().strftime("%Y년 %m월 %d일")

        def _or_none(val: str, label: str) -> str:
            return val.strip() if val.strip() else f"(입력 없음 — {label} 미제공)"

        prompt = MACRO_DEAL_PROMPT.format(
            korea_macro       = _or_none(body.korea_macro,       "한국 매크로 지표"),
            global_macro      = _or_none(body.global_macro,      "글로벌 매크로 지표"),
            market_indicators = _or_none(body.market_indicators,  "시장지표"),
            sector            = _or_none(body.sector,             "산업"),
            company           = _or_none(body.company,            "기업명"),
            dart_result       = _or_none(body.dart_result,        "DART 재무분석"),
            industry_intel    = _or_none(body.industry_intel,     "산업 인텔리전스"),
            report_date       = report_date,
            sector_label      = body.sector.strip() or "미지정",
            company_label     = body.company.strip() or "미지정",
        )
        result = call_claude_cli(prompt, timeout=250)
        return {
            "report": result or "(분석 생성 실패 — Claude CLI 응답 없음)",
            "ok": bool(result),
            "report_date": report_date,
            "sector": body.sector or None,
            "company": body.company or None,
        }
    return await loop.run_in_executor(_pool, _run)


@app.get("/api/macro/indicators")
async def macro_indicators():
    """지표 카탈로그 반환 (체크박스 UI용)"""
    return {
        "catalog": MACRO_CATALOG,
        "defaults": _DEFAULT_IDS,
    }


@app.post("/api/macro/analyze")
async def macro_analyze(body: MacroAnalyzeReq):
    loop = asyncio.get_event_loop()
    def _run():
        from datetime import date as _d
        # 사용할 지표 ID 결정
        if body.selected_indicators:
            ids = list(dict.fromkeys(body.selected_indicators))  # 중복 제거
            analysis_type = "사용자 선택 지표 기준 분석"
        else:
            if body.scope == "domestic":
                ids = _DEFAULT_IDS["domestic"]
            elif body.scope == "global":
                ids = _DEFAULT_IDS["global"] + _DEFAULT_IDS["market"]
            else:
                ids = _DEFAULT_IDS["domestic"] + _DEFAULT_IDS["global"] + _DEFAULT_IDS["market"]
            ids = list(dict.fromkeys(ids))
            analysis_type = "대표 지표 기준 종합 분석"

        # 데이터 수집
        import json as _j
        indicator_data = {}
        any_mock = False      # API 매핑은 있지만 호출 실패 (키/네트워크 문제)
        any_no_source = False # 무료 실시간 소스 자체가 없는 지표 (kospi 등)
        for iid in ids:
            d = _get_indicator(iid, force=body.force_refresh)
            # 카탈로그에서 name/unit 보강
            meta = {}
            for cat in MACRO_CATALOG.values():
                for item in cat:
                    if item["id"] == iid:
                        meta = item; break
            indicator_data[iid] = {
                "name":  meta.get("name", iid),
                "unit":  meta.get("unit", ""),
                **{k:v for k,v in d.items() if k not in ("from_cache",)},
            }
            src = d.get("source")
            if src == "mock":
                any_mock = True
            elif src == "no_source":
                any_no_source = True

        scope_label = {"domestic":"국내 중심","global":"글로벌 중심","both":"국내 + 글로벌"}.get(body.scope,"국내+글로벌")
        data_date = _d.today().strftime("%Y년 %m월 %d일")

        sector_section = ""
        if body.sector.strip():
            sector_section = (
                f"대상 산업 **{body.sector}** 이 어떤 매크로 변수(금리·환율·원자재·경기·유동성)에 "
                f"가장 민감한지, 현재 국면에서 유리/불리 여부를 3~5 bullet로 서술하세요."
            )
        else:
            sector_section = "(선택 산업 없음 — 이 섹션은 생략합니다)"

        prompt = MACRO_SNAPSHOT_PROMPT.format(
            data_date      = data_date,
            analysis_type  = analysis_type,
            scope_label    = scope_label,
            sector         = body.sector or "미지정",
            indicators_json= _j.dumps(indicator_data, ensure_ascii=False, indent=2),
            sector_section = sector_section,
        )

        if check_claude_cli():
            report = call_claude_cli(prompt, timeout=220)
        else:
            report = None

        if not report:
            # mock 분석 반환
            report = _macro_mock_report(indicator_data, body.scope, body.sector, analysis_type, data_date)

        return {
            "report":        report,
            "indicators":    indicator_data,
            "analysis_type": analysis_type,
            "scope":         body.scope,
            "sector":        body.sector or None,
            "data_date":     data_date,
            "any_mock":      any_mock,
            "any_no_source": any_no_source,
            "ok":            True,
        }
    return await loop.run_in_executor(_pool, _run)


def _macro_mock_report(indicators: dict, scope: str, sector: str, analysis_type: str, data_date: str) -> str:
    """Claude CLI 없을 때 규칙 기반 mock 리포트"""
    kr_rate = indicators.get("kr_rate",{}).get("value")
    us_rate = indicators.get("us_rate",{}).get("value")
    vix     = indicators.get("vix",{}).get("value")
    usdkrw  = indicators.get("usdkrw",{}).get("value")
    kr_cpi  = indicators.get("kr_cpi",{}).get("value")
    us_cpi  = indicators.get("us_cpi",{}).get("value")
    us_10y  = indicators.get("us_10y",{}).get("value")

    phase = "혼조"
    if vix and vix > 30: phase = "Risk-Off"
    elif vix and vix < 18: phase = "Risk-On"

    lines = [
        f"# Macro Snapshot",
        f"> {analysis_type} | {data_date} 기준 (규칙 기반 분석 — Claude CLI 미사용)",
        "",
        f"**현재 국면**: {phase}",
        "",
    ]
    if kr_rate: lines.append(f"- **금리 환경**: 한국 기준금리 {kr_rate}%, {'고금리 환경 지속' if kr_rate>=3.0 else '완화적 환경'}")
    if usdkrw:  lines.append(f"- **환율 환경**: USD/KRW {usdkrw:,.0f}원, {'원화 약세' if usdkrw>=1300 else '원화 강세'} 국면")
    if kr_cpi:  lines.append(f"- **물가 환경**: 한국 CPI {kr_cpi}% (YoY), {'목표 수렴' if kr_cpi<=2.5 else '목표 상회'}")
    if vix:     lines.append(f"- **시장심리**: VIX {vix}pt, {'위험 회피 확대' if vix>25 else '비교적 안정적'}")
    lines += ["","---","","*Claude CLI가 설치되지 않아 규칙 기반 분석을 표시합니다. 전체 분석은 Claude CLI 설치 후 이용 가능합니다.*"]
    return "\n".join(lines)


@app.post("/api/company-industry-trends")
async def company_industry_trends(body: CompanyTrendsReq):
    loop = asyncio.get_event_loop()
    def _run():
        scope     = body.scope or "domestic"
        ko_query  = f"{body.corp_name} {body.industry}".strip()
        # 기업명 기반 글로벌 키워드 (영문 변환 포함)
        en_kws = _build_global_kws(body.corp_name, body.industry) if body.corp_name else _get_en_keywords(body.industry)
        # 국내 뉴스
        news_ko: list = search_naver_news(ko_query, display=15) if scope != "global_first" else []
        # 글로벌 뉴스
        news_en: list = search_global_news(en_kws, display_per_kw=5) if scope in ("global", "global_first") else []
        all_news  = news_ko + news_en
        if news_en:
            all_news = translate_articles_if_needed(all_news)
        has_data  = bool(all_news) or bool(body.text_input.strip())
        fin_table = _build_fin_table(body.corp_name, body.yearly) if body.yearly else ""
        result    = generate_company_trends_claude(body.corp_name, all_news, body.text_input, fin_table)
        if not result:
            result = _company_trends_mock(body.corp_name, body.industry, has_data)
        else:
            result = _normalize_company_result(result, all_news)
        deal_highlights = extract_deal_highlights(all_news)
        deal_highlights = generate_deal_highlights_claude(deal_highlights, body.industry or body.corp_name)
        return {
            **result,
            "news": all_news,
            "deal_highlights": deal_highlights,
            "deal_highlights_count": len(deal_highlights),
            "news_ko_count": len(news_ko),
            "news_en_count": len(news_en),
            "naver_available": bool(NAVER_CLIENT_ID),
            "global_available": bool(NEWS_API_KEY) or True,
            "scope": scope,
        }
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/industry-analysis")
async def industry_analysis(body: IndustryAnalysisReq):
    loop = asyncio.get_event_loop()
    def _run():
        scope    = body.scope or "domestic"
        en_kws   = _get_en_keywords(body.industry)
        news_ko  = search_naver_news(body.industry, display=15) if scope != "global_first" else []
        news_en  = search_global_news(en_kws, display_per_kw=5) if scope in ("global", "global_first") else []
        all_news = news_ko + news_en
        if news_en:
            all_news = translate_articles_if_needed(all_news)
        has_data = bool(all_news) or bool(body.text_input.strip())
        result   = generate_industry_analysis_claude(body.industry, all_news, body.text_input)
        if not result:
            result = _industry_analysis_mock(body.industry, has_data)
        else:
            result = _normalize_industry_result(result, all_news)
        deal_highlights = extract_deal_highlights(all_news)
        deal_highlights = generate_deal_highlights_claude(deal_highlights, body.industry)
        return {
            **result,
            "sources": all_news,
            "deal_highlights": deal_highlights,
            "deal_highlights_count": len(deal_highlights),
            "news_ko_count": len(news_ko),
            "news_en_count": len(news_en),
            "naver_available": bool(NAVER_CLIENT_ID),
            "global_available": bool(NEWS_API_KEY) or True,
            "scope": scope,
        }
    return await loop.run_in_executor(_pool, _run)


# ── Insights CRUD ──────────────────────────────────────────────────────────────
class InsightIn(BaseModel):
    created_at:   str = ""
    author:       str = "이기준"
    company_name: str = ""
    sector:       str = ""
    source_type:  str = "직접 입력"
    source_title: str = ""
    source_url:   str = ""
    key_content:  str = ""
    user_insight: str = ""
    impact_level: str = "Medium"
    next_actions: str = ""
    tags:         str = ""

def _row_to_dict(row) -> dict:
    return dict(row)

@app.post("/api/insights")
async def create_insight(body: InsightIn):
    if not body.created_at:
        body.created_at = datetime.now().strftime("%Y-%m-%d")
    with _db() as conn:
        cur = conn.execute(
            """INSERT INTO insights
               (created_at,author,company_name,sector,source_type,source_title,
                source_url,key_content,user_insight,impact_level,next_actions,tags)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (body.created_at, body.author, body.company_name, body.sector,
             body.source_type, body.source_title, body.source_url,
             body.key_content, body.user_insight, body.impact_level,
             body.next_actions, body.tags)
        )
        conn.commit()
        new_id = cur.lastrowid
    with _db() as conn:
        row = conn.execute("SELECT * FROM insights WHERE id=?", (new_id,)).fetchone()
    return _row_to_dict(row)

@app.get("/api/insights")
async def list_insights(
    company: Optional[str] = None,
    sector:  Optional[str] = None,
    impact:  Optional[str] = None,
    tag:     Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
):
    query = "SELECT * FROM insights WHERE 1=1"
    params = []
    if company:
        query += " AND company_name LIKE ?"; params.append(f"%{company}%")
    if sector:
        query += " AND sector LIKE ?";       params.append(f"%{sector}%")
    if impact:
        query += " AND impact_level=?";      params.append(impact)
    if tag:
        query += " AND tags LIKE ?";         params.append(f"%{tag}%")
    if date_from:
        query += " AND created_at >= ?";     params.append(date_from)
    if date_to:
        query += " AND created_at <= ?";     params.append(date_to)
    query += " ORDER BY id DESC"
    with _db() as conn:
        rows = conn.execute(query, params).fetchall()
    return [_row_to_dict(r) for r in rows]

@app.get("/api/insights/{insight_id}")
async def get_insight(insight_id: int):
    with _db() as conn:
        row = conn.execute("SELECT * FROM insights WHERE id=?", (insight_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    return _row_to_dict(row)

@app.put("/api/insights/{insight_id}")
async def update_insight(insight_id: int, body: InsightIn):
    with _db() as conn:
        row = conn.execute("SELECT id FROM insights WHERE id=?", (insight_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Not found")
        conn.execute(
            """UPDATE insights SET
               created_at=?,author=?,company_name=?,sector=?,source_type=?,
               source_title=?,source_url=?,key_content=?,user_insight=?,
               impact_level=?,next_actions=?,tags=?
               WHERE id=?""",
            (body.created_at, body.author, body.company_name, body.sector,
             body.source_type, body.source_title, body.source_url,
             body.key_content, body.user_insight, body.impact_level,
             body.next_actions, body.tags, insight_id)
        )
        conn.commit()
        row = conn.execute("SELECT * FROM insights WHERE id=?", (insight_id,)).fetchone()
    return _row_to_dict(row)

@app.delete("/api/insights/{insight_id}")
async def delete_insight(insight_id: int):
    with _db() as conn:
        row = conn.execute("SELECT id FROM insights WHERE id=?", (insight_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Not found")
        conn.execute("DELETE FROM insights WHERE id=?", (insight_id,))
        conn.commit()
    return {"ok": True}


@app.get("/api/health")
async def health():
    return {"status": "ok", "dart_ready": _dart_ready, "claude_cli": check_claude_cli()}



# ═══════════════════════════════════════════════════════════════════════════════
# 매크로 뉴스 브리프
# ═══════════════════════════════════════════════════════════════════════════════

_MNB_TOPIC_MAP: dict = {
    "interest_rate": {
        "ko": ["한국은행 기준금리", "금리 인상 인하", "통화정책"],
        "en": ["Fed rate decision", "interest rate hike", "monetary policy FOMC"],
    },
    "fx": {
        "ko": ["원달러 환율", "달러 강세 약세"],
        "en": ["dollar index DXY", "USD exchange rate", "currency"],
    },
    "inflation": {
        "ko": ["국내 물가", "한국 CPI 원자재"],
        "en": ["US CPI inflation", "oil price commodity"],
    },
    "growth": {
        "ko": ["한국 GDP 수출 둔화 소비심리"],
        "en": ["US GDP global recession economic growth"],
    },
    "employment": {
        "ko": ["고용 실업률 노동시장"],
        "en": ["US jobs report unemployment nonfarm payroll"],
    },
    "liquidity": {
        "ko": ["유동성 자금시장"],
        "en": ["liquidity credit conditions money market"],
    },
    "credit": {
        "ko": ["부동산 PF 회사채 크레딧"],
        "en": ["credit spread corporate bond high yield real estate"],
    },
    "equity_market": {
        "ko": ["코스피 코스닥 금융시장 증시"],
        "en": ["stock market volatility equity S&P 500"],
    },
    "geopolitics": {
        "ko": ["지정학 공급망 무역 갈등"],
        "en": ["geopolitical risk supply chain tariff trade war"],
    },
    "policy": {
        "ko": ["정책 규제 법안"],
        "en": ["government policy regulation legislation"],
    },
}

_MNB_DEFAULT_KO: list = [
    "한국 금리", "한국은행 기준금리", "원달러 환율", "국내 물가",
    "한국 GDP", "수출 둔화", "소비심리", "부동산 PF", "코스피", "금융시장",
]

_MNB_DEFAULT_EN: list = [
    "Fed rate", "US Treasury yield", "US CPI", "dollar index",
    "oil price", "global recession", "US jobs report", "geopolitical risk",
]

_MNB_CLASSIFY_KW: dict = {
    "interest_rate": ["금리", "기준금리", "rate", "fed", "fomc", "통화정책", "monetary", "boj", "ecb"],
    "fx":            ["환율", "달러", "원달러", "dollar", "dxy", "exchange", "currency", "yen", "yuan"],
    "inflation":     ["물가", "cpi", "ppi", "inflation", "원자재", "유가", "oil", "commodity", "금값"],
    "growth":        ["gdp", "성장", "경기", "소비", "수출", "recession", "growth", "trade", "retail"],
    "employment":    ["고용", "실업", "jobs", "unemployment", "labor", "payroll", "고용률"],
    "liquidity":     ["유동성", "자금", "liquidity", "repo", "credit condition", "머니"],
    "credit":        ["부동산", "pf", "회사채", "크레딧", "credit", "bond", "high yield", "clo", "리파이"],
    "equity_market": ["증시", "코스피", "코스닥", "주가", "stock", "equity", "s&p", "vix", "나스닥"],
    "geopolitics":   ["지정학", "공급망", "tariff", "geopolit", "supply chain", "war", "sanction", "무역"],
    "policy":        ["정책", "규제", "정부", "policy", "regulation", "government", "법안"],
}

_MNB_TOPIC_LABELS: dict = {
    "interest_rate": "금리/통화정책",
    "fx":            "환율/달러",
    "inflation":     "물가/원자재",
    "growth":        "경기/GDP/소비",
    "employment":    "고용",
    "liquidity":     "유동성/자금시장",
    "credit":        "부동산PF/크레딧",
    "equity_market": "증시/위험자산",
    "geopolitics":   "지정학/공급망",
    "policy":        "정책/규제",
    "other":         "기타",
}

_MNB_HIGH_IMPORTANCE_SRCS = {
    "연합뉴스", "연합인포맥스", "한국경제", "매일경제", "조선일보",
    "bloomberg", "reuters", "ft.com", "wsj", "financial times", "cnbc",
    "economist", "nytimes", "ap news", "marketwatch",
}


def _mnb_classify_topic(title: str, desc: str = "") -> list:
    text = f"{title} {desc}".lower()
    found = [t for t, kws in _MNB_CLASSIFY_KW.items() if any(kw in text for kw in kws)]
    return found if found else ["other"]


def _mnb_score_importance(article: dict) -> int:
    source = article.get("source", "").lower()
    text   = f"{article.get('title','')} {article.get('description','')}".lower()
    topics = article.get("macro_topics", [])
    score  = 2
    if any(s in source for s in _MNB_HIGH_IMPORTANCE_SRCS):
        score += 1
    if len(topics) >= 3:
        score += 1
    impact_kws = ["급등", "급락", "충격", "비상", "위기", "collapse", "crash", "shock", "surge", "plunge"]
    if any(k in text for k in impact_kws):
        score += 1
    return min(score, 5)


def _build_macro_news_queries(scope: str, topics: list, extra: list) -> tuple:
    ko_q, en_q = [], []
    if topics:
        for t in topics:
            m = _MNB_TOPIC_MAP.get(t, {})
            ko_q.extend(m.get("ko", [])[:2])
            en_q.extend(m.get("en", [])[:2])
    else:
        ko_q = _MNB_DEFAULT_KO[:]
        en_q = _MNB_DEFAULT_EN[:]
    for kw in extra:
        if any('가' <= c <= '힣' for c in kw):
            ko_q.append(kw)
        else:
            en_q.append(kw)
    ko_q = list(dict.fromkeys(ko_q))[:7]
    en_q = list(dict.fromkeys(en_q))[:7]
    return ko_q, en_q


def _mnb_date_from(date_range: str) -> str:
    from datetime import date as _d, timedelta as _td
    offsets = {"1d": 1, "1w": 7, "1m": 30, "3m": 90}
    return (_d.today() - _td(days=offsets.get(date_range, 7))).isoformat()


def _mnb_search_newsapi(query: str, n: int, from_date: str) -> list:
    if not NEWS_API_KEY:
        return []
    try:
        r = requests.get(
            "https://newsapi.org/v2/everything",
            params={"q": query, "language": "en", "sortBy": "publishedAt",
                    "pageSize": min(n, 20), "apiKey": NEWS_API_KEY, "from": from_date},
            timeout=10,
        )
        if r.status_code != 200:
            return []
        return [
            {"title": a.get("title",""), "description": a.get("description") or "",
             "url": a.get("url",""), "source": a.get("source",{}).get("name",""),
             "published_at": (a.get("publishedAt") or "")[:10], "source_language": "en"}
            for a in r.json().get("articles", [])
            if a.get("title") and "[Removed]" not in a.get("title","")
        ]
    except Exception:
        return []


def _mnb_search_gdelt(query: str, n: int, date_range: str) -> list:
    timespan = {"1d":"1d","1w":"1w","1m":"1m","3m":"3m"}.get(date_range, "1w")
    try:
        r = requests.get(
            "https://api.gdeltproject.org/api/v2/doc/doc",
            params={"query": query, "mode": "artlist", "maxrecords": min(n, 25),
                    "format": "json", "sort": "DateDesc", "timespan": timespan},
            timeout=12,
        )
        if r.status_code != 200:
            return []
        return [
            {"title": a.get("title",""), "description": "",
             "url": a.get("url",""), "source": _extract_domain(a.get("url","")),
             "published_at": (a.get("seendate") or "")[:10], "source_language": "en"}
            for a in (r.json().get("articles") or [])
            if a.get("title") and a.get("url")
        ]
    except Exception:
        return []


def _mnb_cache_get(key: str) -> dict | None:
    with _db() as conn:
        row = conn.execute(
            "SELECT value_json FROM macro_news_cache WHERE cache_key=? "
            "AND datetime(fetched_at) > datetime('now','-6 hours')", (key,)
        ).fetchone()
    if row:
        import json as _j
        return _j.loads(row[0])
    return None


def _mnb_cache_set(key: str, data: dict) -> None:
    import json as _j
    with _db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO macro_news_cache(cache_key,value_json,fetched_at) "
            "VALUES(?,?,datetime('now'))", (key, _j.dumps(data, ensure_ascii=False))
        )


MACRO_NEWS_BRIEF_PROMPT = """\
너는 VC, PE, M&A, IPO, CFO 관점에서 매크로 뉴스와 경제지표를 해석하는 Macro Deal Intelligence Analyst다.

## 분석 기준일
{data_date}

## 분석 범위 / 기간
{scope_label} / {date_range_label}
대상 산업: {sector}

## 정량 지표 (ECOS/FRED)
{macro_data}

## 수집된 뉴스 (국내 {ko_count}건 / 글로벌 {en_count}건)

### 국내 뉴스
{ko_news}

### 글로벌 뉴스
{en_news}

---

## 분석 지침
- ECOS/FRED 정량 지표와 최근 뉴스 검색 결과(제목/요약 기반)를 함께 해석하라.
- 반복적으로 등장하는 시장 이슈를 우선순위화하라.
- 최종 결과는 한국어 마크다운으로 작성하라.
- 확인되지 않은 수치는 단정하지 말라. 투자 추천 표현은 피하라.
- VC, PE, M&A, IPO, CFO 관점의 시사점을 포함하라.

## 출력 형식

# Macro News Brief

## 1. Executive Summary
[3~5 bullet — 핵심 매크로 동향 요약]

## 2. Market Moving Issues
[지금 가장 중요한 시장 이슈 3~5개, 각 1~2줄]

## 3. 국내 매크로 동향
[국내 뉴스 기반 동향, 3~5 bullet]

## 4. 글로벌 매크로 동향
[글로벌 뉴스 기반 동향, 3~5 bullet]

## 5. 금리 & 유동성
[금리/통화정책/유동성 해석, 2~3 bullet]

## 6. 환율 & 달러
[환율/달러 해석, 2~3 bullet]

## 7. 물가 & 원자재
[물가/유가/상품 해석, 2~3 bullet]

## 8. 경기 & 수요
[GDP/소비/무역 해석, 2~3 bullet]

## 9. 크레딧 & 리스크 센티먼트
[크레딧/부동산PF/회사채/증시 해석, 2~3 bullet]

## 10. Deal Implication
[VC, PE, M&A, IPO, CFO 관점 시사점, 4~6 bullet]

## 11. CFO Checkpoints
[CFO가 지금 점검할 사항, 3~5 bullet]

## 12. Due Diligence Questions
[현 매크로 환경에서 실사 확인 질문, 3~5개]
"""


def _build_mnb_prompt(body, ko_arts: list, en_arts: list, macro_data_str: str) -> str:
    from datetime import date as _d
    scope_label     = {"domestic":"국내 중심","global":"글로벌 중심","both":"국내+글로벌"}.get(body.scope,"국내+글로벌")
    date_range_label = {"1d":"최근 1일","1w":"최근 1주","1m":"최근 1개월","3m":"최근 3개월"}.get(body.date_range,"최근 1주")

    def _fmt_arts(arts, n=12):
        rows = []
        for i, a in enumerate(arts[:n], 1):
            line = f"{i}. [{a.get('source','')}] {a.get('title','')} ({a.get('published_at','')})"
            if a.get("description"):
                line += f"\n   {a['description'][:130]}"
            rows.append(line)
        return "\n".join(rows) if rows else "없음"

    return MACRO_NEWS_BRIEF_PROMPT.format(
        data_date        = _d.today().strftime("%Y년 %m월 %d일"),
        scope_label      = scope_label,
        date_range_label = date_range_label,
        sector           = body.sector or "미지정",
        macro_data       = macro_data_str or "정량 지표 미포함",
        ko_news          = _fmt_arts(ko_arts),
        en_news          = _fmt_arts(en_arts),
        ko_count         = len(ko_arts),
        en_count         = len(en_arts),
    )


def _mock_mnb_report(body, ko_arts: list, en_arts: list) -> str:
    scope_label     = {"domestic":"국내 중심","global":"글로벌 중심","both":"국내+글로벌"}.get(body.scope,"국내+글로벌")
    date_range_label = {"1d":"최근 1일","1w":"최근 1주","1m":"최근 1개월","3m":"최근 3개월"}.get(body.date_range,"최근 1주")
    all_arts = ko_arts + en_arts
    topics_found = sorted(set(t for a in all_arts for t in a.get("macro_topics",[])))

    if not all_arts:
        return f"""# Macro News Brief

## 1. Executive Summary
- 수집된 뉴스 없음 — .env에 NAVER_CLIENT_ID/SECRET 또는 NEWS_API_KEY 설정 시 실시간 수집
- 분석 범위: {scope_label} / {date_range_label}
- GDELT API는 별도 키 없이 사용 가능하나 응답 속도가 느릴 수 있음

## 10. Deal Implication
- API 키 설정 후 실시간 뉴스 기반 딜 시사점 분석 가능

## 11. CFO Checkpoints
- 금리 환경 점검 (지표 분석 탭 참조)
- 환율 리스크 및 외화 익스포저 확인
- 크레딧 시장 동향 및 리파이낸싱 일정 점검

## 12. Due Diligence Questions
- 주요 고객/공급망의 지정학 노출도는?
- 금리 상승 시나리오에서 이자보상배율 변화는?
- 원화 약세 시 외화 부채 비중 및 헤징 현황은?
"""

    topic_str = ", ".join([_MNB_TOPIC_LABELS.get(t, t) for t in topics_found[:5]]) or "분류 중"
    ko_bullets = "\n".join([f"- [{a.get('source','')}] {a.get('title','')[:60]}" for a in ko_arts[:4]])
    en_bullets = "\n".join([f"- [{a.get('source','')}] {a.get('title','')[:70]}" for a in en_arts[:4]])

    return f"""# Macro News Brief

## 1. Executive Summary
- 국내 뉴스 {len(ko_arts)}건 / 글로벌 뉴스 {len(en_arts)}건 수집 ({scope_label} / {date_range_label})
- 주요 매크로 주제: {topic_str}
- Claude CLI 연동 시 AI 기반 심층 분석이 자동 생성됩니다

## 2. Market Moving Issues
{ko_bullets}
{en_bullets}

## 3. 국내 매크로 동향
{"".join([f"- [{a.get('source','')}] {a.get('title','')}" + chr(10) for a in ko_arts[:5]]) or "- 국내 뉴스 없음"}

## 4. 글로벌 매크로 동향
{"".join([f"- [{a.get('source','')}] {a.get('title','')}" + chr(10) for a in en_arts[:5]]) or "- 글로벌 뉴스 없음"}

## 10. Deal Implication
- 뉴스 주제 분포: **{topic_str}** 이슈 부각
- Claude CLI 연동 시 VC/PE/M&A/IPO/CFO 관점 시사점 자동 분석

## 11. CFO Checkpoints
- 금리 및 자금조달 비용 변화 점검
- 환율 리스크 및 외화 익스포저 확인
- 크레딧 시장 동향 및 리파이낸싱 일정 점검

## 12. Due Diligence Questions
- 주요 고객/공급망의 지정학 노출도는?
- 금리 상승 시나리오에서 이자보상배율 변화는?
- 원화 약세 시 외화 부채 비중 및 헤징 현황은?
"""


class MacroNewsBriefReq(BaseModel):
    scope:               str  = "both"
    topics:              list = []
    date_range:          str  = "1w"
    keywords:            list = []
    max_results:         int  = 20
    macro_data:          dict = {}
    sector:              str  = ""
    company_name:        str  = ""
    force_refresh:       bool = False


@app.post("/api/macro/news-brief")
async def macro_news_brief(body: MacroNewsBriefReq):
    loop = asyncio.get_event_loop()

    def _run():
        import json as _j, hashlib as _hs
        from concurrent.futures import ThreadPoolExecutor, as_completed

        cache_key = _hs.md5(_j.dumps({
            "scope": body.scope, "topics": sorted(body.topics),
            "date_range": body.date_range, "keywords": sorted(body.keywords),
            "max_results": body.max_results,
        }, sort_keys=True).encode()).hexdigest()

        if not body.force_refresh:
            cached = _mnb_cache_get(cache_key)
            if cached:
                return {**cached, "from_cache": True, "ok": True}

        ko_queries, en_queries = _build_macro_news_queries(body.scope, body.topics, body.keywords)
        from_date = _mnb_date_from(body.date_range)
        seen_urls: set = set()
        ko_articles: list = []
        en_articles: list = []

        # Domestic news
        if body.scope in ("domestic", "both") and NAVER_CLIENT_ID:
            per_q = max(3, body.max_results // max(len(ko_queries[:5]), 1))
            for q in ko_queries[:5]:
                for a in search_naver_news(q, display=per_q):
                    url = a.get("url","")
                    if url and url not in seen_urls:
                        seen_urls.add(url); ko_articles.append(a)

        # Global news (parallel)
        if body.scope in ("global", "both"):
            def _fetch_en(q):
                arts = _mnb_search_newsapi(q, 5, from_date)
                if not arts:
                    arts = _mnb_search_gdelt(q, 5, body.date_range)
                return arts
            with ThreadPoolExecutor(max_workers=4) as ex:
                futs = [ex.submit(_fetch_en, q) for q in en_queries[:5]]
                for fut in as_completed(futs, timeout=60):
                    try:
                        for a in (fut.result() or []):
                            url = a.get("url","")
                            if url and url not in seen_urls:
                                seen_urls.add(url); en_articles.append(a)
                    except Exception:
                        pass

        # Classify + score
        limit = max(body.max_results // 2, 8)
        for a in ko_articles[:limit] + en_articles[:limit]:
            a["macro_topics"]     = _mnb_classify_topic(a.get("title",""), a.get("description",""))
            a["importance_score"] = _mnb_score_importance(a)

        ko_articles = sorted(ko_articles[:limit], key=lambda x: x.get("importance_score",2), reverse=True)
        en_articles = sorted(en_articles[:limit], key=lambda x: x.get("importance_score",2), reverse=True)
        # B-1 en_articles fallback from ko (2026-05-21)
        if not en_articles and ko_articles:
            # NewsAPI rate-limit / GDELT 빈 응답 시 fallback —
            # ko_articles 안 외신 source (영문 / 비한국어) 를 분리.
            foreign = [
                a for a in ko_articles
                if (a.get("source_language") or "ko") != "ko"
                or any(s in (a.get("source") or "").lower() for s in
                       ("coindesk", "reuters", "bloomberg", "cnbc", "ft.com",
                        "wsj", "nytimes", "sabah", "portfolio", "cfi"))
            ]
            if foreign:
                en_articles = sorted(foreign[:limit],
                                     key=lambda x: x.get("importance_score", 2),
                                     reverse=True)
                # ko_articles 에서 외신 항목 제거 (중복 방지)
                _foreign_urls = {a.get("url") for a in foreign}
                ko_articles = [a for a in ko_articles if a.get("url") not in _foreign_urls]
        all_articles = (ko_articles + en_articles)[:body.max_results]

        # Claude prompt
        macro_data_str = _j.dumps(body.macro_data, ensure_ascii=False, indent=2) if body.macro_data else ""
        prompt = _build_mnb_prompt(body, ko_articles, en_articles, macro_data_str)

        report = None
        if check_claude_cli():
            report = call_claude_cli(prompt, timeout=230)
        if not report:
            report = _mock_mnb_report(body, ko_articles, en_articles)

        result = {
            "ok":            True,
            "scope":         body.scope,
            "date_range":    body.date_range,
            "report":        report,
            "articles":      all_articles,
            "ko_articles":   ko_articles,
            "en_articles":   en_articles,
            "article_count": {"ko": len(ko_articles), "en": len(en_articles)},
            "topics_found":  sorted(set(t for a in all_articles for t in a.get("macro_topics",[]))),
            "has_real_news": len(all_articles) > 0,
            "sector":        body.sector or None,
            "from_cache":    False,
        }
        _mnb_cache_set(cache_key, result)
        return result

    return await loop.run_in_executor(_pool, _run)




# ═══════════════════════════════════════════════════════════════════════════════
# 백업 / 내보내기
# ═══════════════════════════════════════════════════════════════════════════════

from fastapi import Query as _Q
from fastapi.responses import Response as _Resp


def _all_insights() -> list:
    with _db() as conn:
        rows = conn.execute("SELECT * FROM insights ORDER BY created_at DESC").fetchall()
    return [dict(r) for r in rows]


@app.post("/api/backup/db")
async def backup_db():
    """SQLite DB를 backups/ 폴더에 타임스탬프 파일명으로 복사."""
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = BACKUP_DIR / f"standard_view_backup_{ts}.db"
        shutil.copy2(str(DB_PATH), str(dest))
        size_kb = round(dest.stat().st_size / 1024, 1)
        return {"ok": True, "filename": dest.name, "size_kb": size_kb}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/backup/list")
async def list_backups():
    """backups/ 폴더의 .db 파일 목록 반환."""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(BACKUP_DIR.glob("standard_view_backup_*.db"), reverse=True)
    return [
        {
            "filename": f.name,
            "size_kb":  round(f.stat().st_size / 1024, 1),
            "created":  datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        }
        for f in files[:30]
    ]


@app.get("/api/backup/export")
async def export_insights(format: str = _Q("json", pattern="^(json|csv|md)$")):
    """인사이트를 JSON / CSV / Markdown으로 내보내기."""
    items = _all_insights()
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format == "json":
        body  = json.dumps(items, ensure_ascii=False, indent=2)
        fname = f"insights_{ts}.json"
        return _Resp(
            content=body.encode("utf-8"),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )

    if format == "csv":
        headers = ["id","created_at","author","company_name","sector","source_type",
                   "source_title","source_url","key_content","user_insight",
                   "impact_level","next_actions","tags"]
        def _cell(v):
            return '"' + str(v or "").replace('"', '""') + '"'
        rows = [",".join(headers)] + [",".join(_cell(r.get(h)) for h in headers) for r in items]
        body  = "﻿" + "\n".join(rows)   # BOM for Excel
        fname = f"insights_{ts}.csv"
        return _Resp(
            content=body.encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )

    # Markdown
    lines = ["# 인사이트 로그", f"exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ""]
    for it in items:
        tags = " ".join(f"#{t.strip()}" for t in (it.get("tags") or "").split(",") if t.strip())
        lines += [
            f"## [{it.get('impact_level','?')}] {it.get('source_title') or it.get('company_name') or '(제목 없음)'}",
            f"- **날짜**: {it.get('created_at','')}  **기록자**: {it.get('author','')}",
            f"- **기업**: {it.get('company_name') or '-'}  **산업**: {it.get('sector') or '-'}",
            f"- **출처**: {it.get('source_type','')}",
        ]
        if it.get("source_url"):
            lines.append(f"- **URL**: {it['source_url']}")
        lines += ["", f"**주요 내용**\n{it.get('key_content','')}"]
        if it.get("user_insight"):
            lines += ["", f"**내 인사이트**\n{it['user_insight']}"]
        if it.get("next_actions"):
            lines += ["", f"**후속 확인사항**\n{it['next_actions']}"]
        if tags:
            lines.append(f"\n{tags}")
        lines += ["", "---", ""]
    body  = "\n".join(lines)
    fname = f"insights_{ts}.md"
    return _Resp(
        content=body.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )



if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8002, reload=False)





# ─── Phase B-2 (2026-05-19): NOAH analysis ingest ─────────────────
# NOAH stock-bot 가 분석 완료 시 POST → JSONL append.
# daily_generator 가 다음 08:00 KST 에 읽어서 HTML 새 section inject.

import os as _os_b2
import json as _json_b2
from datetime import date as _date_b2, datetime as _dt_b2

_NOAH_JSONL = Path(_os_b2.path.expanduser("~/standardview/noah_analyses.jsonl"))


@app.post("/api/noah/analysis")
async def noah_analysis_ingest(payload: dict):
    if not isinstance(payload, dict) or not payload.get("ticker"):
        return {"ok": False, "error": "ticker required"}
    record = dict(payload)
    now = _dt_b2.now()
    record["ts"] = now.isoformat(timespec="seconds")
    record["date"] = now.date().isoformat()
    try:
        _NOAH_JSONL.parent.mkdir(parents=True, exist_ok=True)
        with _NOAH_JSONL.open("a", encoding="utf-8") as f:
            f.write(_json_b2.dumps(record, ensure_ascii=False) + "\n")
        return {"ok": True, "stored": record["ts"]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/noah/today")
async def noah_today(date: str | None = None):
    target = date or _date_b2.today().isoformat()
    if not _NOAH_JSONL.exists():
        return {"ok": True, "date": target, "analyses": []}
    out = []
    try:
        with _NOAH_JSONL.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = _json_b2.loads(line)
                except Exception:
                    continue
                if rec.get("date") == target:
                    out.append(rec)
        return {"ok": True, "date": target, "analyses": out}
    except Exception as e:
        return {"ok": False, "error": str(e), "date": target, "analyses": []}




# ─── /dashboard route — serve latest daily brief (always-fresh) ───
# HTTP Basic Auth required. Credentials from .env (SV_DASHBOARD_USER /
# SV_DASHBOARD_PW). 모바일 브라우저 (iOS Safari / Chrome / Firefox /
# Brave 모두) Basic Auth popup 지원. 다만 iOS Chrome 일부 버전이 raw
# IP HTTP Basic Auth popup 띄우지 않는 케이스 — Safari 또는 Firefox
# 시도 권장 (NOAH dashboard 와 동일 양상).
from fastapi import Depends as _Dep_dash, HTTPException as _HE_dash
from fastapi.security import HTTPBasic as _HB_dash, HTTPBasicCredentials as _HBC_dash
from fastapi.responses import FileResponse as _FileResp_dash
from pathlib import Path as _Path_dash
import secrets as _sec_dash
import os as _os_dash

# ── SV theme injection ────────────────────────────────────────────────────────
# Light/dark auto-switch: 19:00–07:00 KST → dark, rest → light.
# Same timing as NOAH dashboard. Injected into every SV HTML page at serve time
# so the base.html template stays untouched.
_SV_THEME_CSS_LIGHT = """\
<style>
:root[data-theme="light"]{
  --bg:#f8fafc;--text:#1f2937;--text-muted:#6b7280;
  --accent:#2563eb;--accent-soft:rgba(37,99,235,0.1);
  --border:#e2e8f0;--surface:rgba(255,255,255,0.92);
  --surface-2:rgba(248,250,252,0.9);--cyan:#0891b2;--muted:#6b7280;
}
:root[data-theme="light"] details{background:rgba(255,255,255,0.85)}
:root[data-theme="light"] .badge-daily{background:rgba(37,99,235,0.12);color:#1d4ed8}
:root[data-theme="light"] .badge-weekly{background:rgba(5,150,105,0.12);color:#065f46}
:root[data-theme="light"] .badge-sectors{background:rgba(180,83,9,0.12);color:#92400e}
</style>"""

_SV_THEME_JS = """\
<script>
(function(){
  function apply(){
    var h=parseInt(new Intl.DateTimeFormat('en-US',{
      timeZone:'Asia/Seoul',hour:'numeric',hour12:false
    }).format(new Date()),10)%24;
    document.documentElement.dataset.theme=(h>=19||h<7)?'dark':'light';
  }
  apply();setInterval(apply,60000);
})();
</script>"""


def _sv_inject_theme(html: str) -> str:
    """Inject light/dark theme CSS + JS — case-insensitive head/body search."""
    inject = _SV_THEME_CSS_LIGHT + "\n" + _SV_THEME_JS + "\n"
    lower = html.lower()
    # 1st choice: before </head> (any case)
    idx = lower.find("</head>")
    if idx >= 0:
        return html[:idx] + inject + html[idx:]
    # 2nd choice: right after opening <body ...> tag
    body_start = lower.find("<body")
    if body_start >= 0:
        body_end = html.find(">", body_start)
        return html[:body_end + 1] + "\n" + inject + html[body_end + 1:]
    # Last resort: prepend
    return inject + html

# ─────────────────────────────────────────────────────────────────────────────

_dash_security = _HB_dash()
_SV_USER = _os_dash.environ.get("SV_DASHBOARD_USER", "")
_SV_PW   = _os_dash.environ.get("SV_DASHBOARD_PW", "")


def _verify_dashboard_auth(
    credentials: _HBC_dash = _Dep_dash(_dash_security),
) -> str:
    if not (_SV_USER and _SV_PW):
        # Env vars missing — auth disabled (dev fallback). In production
        # set SV_DASHBOARD_USER + SV_DASHBOARD_PW in ~/standardview/.env.
        return "anonymous"
    user_ok = _sec_dash.compare_digest(credentials.username, _SV_USER)
    pw_ok = _sec_dash.compare_digest(credentials.password, _SV_PW)
    if not (user_ok and pw_ok):
        raise _HE_dash(
            status_code=401,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


@app.get("/dashboard")
async def dashboard_latest(_: str = _Dep_dash(_verify_dashboard_auth)):
    latest = _Path_dash.home() / "standardview" / "reports" / "generated" / "latest.html"
    if latest.exists():
        html = latest.read_text(encoding="utf-8")
        return _HTMLResp_arch(_sv_inject_theme(html))
    return {"ok": False, "error": "dashboard not yet generated; wait for hourly timer"}





# ─── /dashboard/archive — index + /dashboard/view/{filename} — viewer ─
# B-3 (2026-05-20): 일자별 archive 누적 표시. <details> 기반 모바일
# 호환 (JS 불필요). 같은 HTTP Basic Auth 적용.
from fastapi.responses import HTMLResponse as _HTMLResp_arch
from collections import defaultdict as _dd_arch
import re as _re_arch


@app.get("/dashboard/archive", response_class=_HTMLResp_arch)
async def dashboard_archive(_: str = _Dep_dash(_verify_dashboard_auth)):
    """List all archived daily briefs grouped by month/date."""
    base = _Path_dash.home() / "standardview" / "reports" / "generated"
    if not base.exists():
        return _HTMLResp_arch("<html><body>no archives</body></html>")

    # Collect {date}_{HHMM}[_{type}].html files
    entries = []
    for f in sorted(base.glob("*.html"), reverse=True):
        if f.name == "latest.html":
            continue
        m = _re_arch.match(
            r"(\d{4}-\d{2}-\d{2})(?:_(\d{4})(?:_(weekly|sectors))?)?\.html$",
            f.name,
        )
        if not m:
            continue
        date_str = m.group(1)
        hhmm = m.group(2) or "0000"  # legacy {date}.html → 0000
        entry_type = m.group(3) or "daily"  # "weekly", "sectors", or "daily"
        entries.append({
            "date": date_str,
            "hhmm": hhmm,
            "filename": f.name,
            "size_kb": f.stat().st_size // 1024,
            "type": entry_type,
        })

    # Group by month → date → entries
    by_month = _dd_arch(lambda: _dd_arch(list))
    for e in entries:
        month = e["date"][:7]  # YYYY-MM
        by_month[month][e["date"]].append(e)

    # Render HTML — <details> for accordion. CSS variables from
    # base.html so visual style stays consistent.
    html_parts = ["""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Standard View — Archive</title>
<style>
:root { --bg: #060A14; --surface: rgba(20,28,48,0.6); --border: rgba(255,255,255,0.08);
  --text: #E8ECF4; --text-muted: #94A3B8; --accent: #3B82F6; }
body { background: var(--bg); color: var(--text); font-family: -apple-system,
  'IBM Plex Sans','Noto Sans KR', sans-serif; padding: 20px; max-width: 800px;
  margin: 0 auto; }
h1 { font-size: 22px; margin-bottom: 24px; }
details { background: var(--surface); border: 1px solid var(--border);
  border-radius: 8px; padding: 12px 16px; margin-bottom: 8px; }
details > summary { cursor: pointer; font-weight: 600; font-size: 15px;
  list-style: none; user-select: none; padding: 4px 0; }
details > summary::after { content: ' ▾'; color: var(--text-muted); float: right; }
details[open] > summary::after { content: ' ▴'; }
details details { margin-left: 16px; margin-top: 8px; background: transparent;
  border-color: rgba(255,255,255,0.04); }
.entry { display: flex; align-items: center; padding: 8px 0; gap: 12px;
  border-top: 1px solid var(--border); }
.entry:first-child { border-top: none; }
.entry-time { font-family: 'IBM Plex Mono', monospace; color: var(--accent);
  font-size: 13px; min-width: 60px; }
.entry-link { color: var(--text); text-decoration: none; flex: 1;
  font-size: 13px; }
.entry-link:hover { color: var(--accent); }
.entry-size { color: var(--text-muted); font-size: 11px; }
.nav { margin-bottom: 16px; }
.nav a { color: var(--accent); text-decoration: none; font-size: 13px; }
.badge { font-size: 10px; padding: 2px 6px; border-radius: 4px; font-weight: 600; }
.badge-daily { background: rgba(59,130,246,0.15); color: #93C5FD; }
.badge-weekly { background: rgba(16,185,129,0.15); color: #6EE7B7; }
.badge-sectors { background: rgba(245,158,11,0.15); color: #FCD34D; }
</style>
<style>
:root[data-theme="light"]{
  --bg:#f8fafc;--text:#1f2937;--text-muted:#6b7280;
  --accent:#2563eb;--border:#e2e8f0;--surface:rgba(255,255,255,0.9);
}
:root[data-theme="light"] details{background:rgba(255,255,255,0.85)}
:root[data-theme="light"] .badge-daily{background:rgba(37,99,235,0.12);color:#1d4ed8}
:root[data-theme="light"] .badge-weekly{background:rgba(5,150,105,0.12);color:#065f46}
:root[data-theme="light"] .badge-sectors{background:rgba(180,83,9,0.12);color:#92400e}
</style>
<script>
(function(){
  function a(){var h=parseInt(new Intl.DateTimeFormat('en-US',{
    timeZone:'Asia/Seoul',hour:'numeric',hour12:false
  }).format(new Date()),10)%24;
  document.documentElement.dataset.theme=(h>=19||h<7)?'dark':'light';}
  a();setInterval(a,60000);
})();
</script>
</head><body>
<div class="nav"><a href="/dashboard">▸ 최신 brief 보기</a></div>
<h1>📚 Standard View — Archive</h1>
"""]

    # Sort months desc (most recent first)
    for month in sorted(by_month.keys(), reverse=True):
        dates_dict = by_month[month]
        total_entries = sum(len(v) for v in dates_dict.values())
        # Outermost <details> for month — open if it's the current month
        from datetime import date as _d
        current_month = _d.today().strftime("%Y-%m")
        is_open = " open" if month == current_month else ""
        html_parts.append(
            f"<details{is_open}><summary>📅 {month} "
            f"<span style='color:var(--text-muted);font-weight:normal'>"
            f"({total_entries}건)</span></summary>"
        )
        for date_str in sorted(dates_dict.keys(), reverse=True):
            day_entries = dates_dict[date_str]
            day_is_open = " open" if date_str == _d.today().isoformat() else ""
            html_parts.append(
                f"<details{day_is_open}><summary>{date_str} "
                f"<span style='color:var(--text-muted);font-weight:normal'>"
                f"({len(day_entries)}건)</span></summary>"
            )
            for e in sorted(day_entries, key=lambda x: x["hhmm"], reverse=True):
                t_label = (
                    f"{e['hhmm'][:2]}:{e['hhmm'][2:]}"
                    if e["hhmm"] != "0000"
                    else "(legacy)"
                )
                _badge_label = {
                    "weekly": "주간브리프", "sectors": "섹터히트맵",
                }.get(e.get("type", "daily"), "브리프")
                _badge_cls = {
                    "weekly": "badge-weekly", "sectors": "badge-sectors",
                }.get(e.get("type", "daily"), "badge-daily")
                html_parts.append(
                    f'<div class="entry">'
                    f'<span class="entry-time">{t_label}</span>'
                    f'<span class="badge {_badge_cls}">{_badge_label}</span>'
                    f'<a class="entry-link" href="/dashboard/view/{e["filename"]}">'
                    f'{e["filename"]}</a>'
                    f'<span class="entry-size">{e["size_kb"]}KB</span>'
                    f'<form method="POST" action="/dashboard/delete/{e["filename"]}"'
                    f' style="display:inline;margin-left:8px"'
                    f'>'
                    f'<button type="submit" title="삭제" style="background:none;'
                    f'border:none;color:var(--text-muted);cursor:pointer;'
                    f'font-size:14px;padding:2px 6px">🗑️</button>'
                    f'</form>'
                    f'</div>'
                )
            html_parts.append("</details>")
        html_parts.append("</details>")

    html_parts.append("</body></html>")
    return _HTMLResp_arch("".join(html_parts))


@app.get("/dashboard/view/{filename}")
async def dashboard_view(
    filename: str,
    _: str = _Dep_dash(_verify_dashboard_auth),
):
    """Serve a specific archived brief. filename must match pattern
    YYYY-MM-DD[_HHMM[_(weekly|sectors)]].html to prevent path traversal."""
    if not _re_arch.match(
        r"^\d{4}-\d{2}-\d{2}(?:_\d{4}(?:_(weekly|sectors))?)?\.html$", filename
    ):
        return {"ok": False, "error": "invalid filename pattern"}
    path = _Path_dash.home() / "standardview" / "reports" / "generated" / filename
    if path.exists():
        html = path.read_text(encoding="utf-8")
        return _HTMLResp_arch(_sv_inject_theme(html))
    return {"ok": False, "error": f"{filename} not found"}


from fastapi.responses import RedirectResponse as _RedirResp_dash


@app.post("/dashboard/delete/{filename}")
async def dashboard_delete(
    filename: str,
    _: str = _Dep_dash(_verify_dashboard_auth),
):
    """Delete archived brief (POST form). latest.html 보호."""
    if filename == "latest.html":
        return {"ok": False, "error": "latest.html protected"}
    if not _re_arch.match(
        r"^\d{4}-\d{2}-\d{2}(?:_\d{4}(?:_(weekly|sectors))?)?\.html$", filename
    ):
        return {"ok": False, "error": "invalid filename pattern"}
    path = _Path_dash.home() / "standardview" / "reports" / "generated" / filename
    if path.exists():
        path.unlink()
    md_path = path.with_suffix(".md")
    if md_path.exists():
        md_path.unlink()
    return _RedirResp_dash("/dashboard/archive", status_code=303)





@app.get("/api/sv-usage/today")
async def sv_usage_today():
    import json as _json, os as _os
    from datetime import date as _d
    path = Path(_os.path.expanduser("~/standardview/sv_usage.jsonl"))
    today = _d.today().isoformat()
    month = today[:7]
    out = {"today_krw": 0.0, "month_krw": 0.0, "today_calls": 0, "month_calls": 0,
           "today_prompt_tok": 0, "today_output_tok": 0}
    if not path.exists():
        return out
    try:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line: continue
                try: rec = _json.loads(line)
                except Exception: continue
                if rec.get("date") == today:
                    out["today_krw"] += rec.get("cost_krw", 0) or 0
                    out["today_calls"] += 1
                    out["today_prompt_tok"] += rec.get("prompt_tok", 0) or 0
                    out["today_output_tok"] += rec.get("output_tok", 0) or 0
                if rec.get("month") == month:
                    out["month_krw"] += rec.get("cost_krw", 0) or 0
                    out["month_calls"] += 1
    except Exception as e:
        out["error"] = str(e)
    out["today_krw"] = round(out["today_krw"], 1)
    out["month_krw"] = round(out["month_krw"], 1)
    return out

# Static frontend dashboard (separate from NOAH). Built by:
#   cd ~/standardview/frontend && npm run build  ->  ~/standardview/dist/
# Mounted LAST so all /api/* routes take precedence.


# sv-telegram-button-endpoint v1
@app.post("/api/push-to-telegram")
def push_to_telegram():
    """Trigger telegram_pusher.py — sends today's brief to the
    STANDARDVIEW channel (MD text + HTML attachment). Used by the
    dashboard's 'Send to Telegram' button as an on-demand push (vs
    the scheduled standardview-daily.timer / hourly.timer)."""
    import subprocess as _sp
    from pathlib import Path as _P
    pusher = _P(__file__).resolve().parent.parent / "scripts" / "telegram_pusher.py"
    venv_py = _P.home() / "standardview" / ".venv" / "bin" / "python"
    if not pusher.exists():
        return {"ok": False, "error": "pusher script missing"}
    py = str(venv_py) if venv_py.exists() else "python3"
    try:
        r = _sp.run([py, str(pusher)], capture_output=True, text=True, timeout=120)
        if r.returncode == 0:
            return {"ok": True, "message": "Telegram push 완료",
                    "stdout": (r.stdout or "")[-300:]}
        return {"ok": False, "error": "pusher failed",
                "stderr": (r.stderr or "")[-300:],
                "stdout": (r.stdout or "")[-300:]}
    except _sp.TimeoutExpired:
        return {"ok": False, "error": "timeout (120s)"}

try:
    from fastapi.staticfiles import StaticFiles
    _DIST_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "dist")
    if os.path.isdir(_DIST_DIR):
        app.mount("/", StaticFiles(directory=_DIST_DIR, html=True), name="dashboard")
except Exception:
    pass

